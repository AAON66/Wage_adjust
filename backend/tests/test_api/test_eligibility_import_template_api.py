"""Phase 32-04 Task 2: GET /eligibility-import/templates/{import_type} 端到端测试。

覆盖（IMPORT-02）：
  - 4 类资格 import_type 模板下载 → 200 + Content-Disposition + Content-Type
  - 模板字节用 openpyxl 可读回，第一行表头匹配 D-02/D-03 字段
  - 工号列文本格式（防 leading zero loss）
  - 未知 import_type → 400
  - employee 角色 → 403
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook


@pytest.mark.parametrize(
    'import_type,expected_headers',
    [
        ('performance_grades', ['员工工号', '年度', '绩效等级']),
        (
            'salary_adjustments',
            ['员工工号', '调薪日期', '调薪类型', '调薪金额'],
        ),
        ('hire_info', ['员工工号', '入职日期', '末次调薪日期']),
        (
            'non_statutory_leave',
            ['员工工号', '年度', '假期天数', '假期类型'],
        ),
    ],
)
def test_template_returns_non_empty_xlsx(
    client_hrbp, import_type, expected_headers,
):
    """4 类资格 import_type 各自能下载非空 .xlsx，表头与 D-02/D-03 一致。"""
    resp = client_hrbp.get(
        f'/api/v1/eligibility-import/templates/{import_type}?format=xlsx'
    )
    assert resp.status_code == 200, resp.text

    # Content-Disposition + Content-Type
    cd = resp.headers.get('content-disposition', '').lower()
    assert 'attachment' in cd
    ct = resp.headers.get('content-type', '')
    assert 'spreadsheetml' in ct or 'octet-stream' in ct

    # 内容非空
    assert len(resp.content) > 100

    # openpyxl 可读回 + 表头校验
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    actual_headers = [c.value for c in ws[1]]
    assert actual_headers == expected_headers

    # 工号列文本格式（防 leading zero loss / EMPNO 前导零修复）
    assert ws.cell(row=2, column=1).number_format == '@'


def test_template_unknown_type_returns_400(client_hrbp):
    """未知 import_type → 400。"""
    resp = client_hrbp.get(
        '/api/v1/eligibility-import/templates/unknown_type?format=xlsx'
    )
    assert resp.status_code == 400


def test_template_employee_role_forbidden(client_employee):
    """employee 角色下载模板 → 403（T-32-04）。"""
    resp = client_employee.get(
        '/api/v1/eligibility-import/templates/hire_info?format=xlsx'
    )
    assert resp.status_code == 403
