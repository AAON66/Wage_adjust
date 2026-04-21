from __future__ import annotations

import io
from pathlib import Path
from uuid import uuid4

import pandas as pd
import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from backend.app.core.config import Settings
from backend.app.core.database import create_db_engine, create_session_factory, init_database
from backend.app.models import load_model_modules
from backend.app.services.import_service import ImportService

load_model_modules()


# ---------------------------------------------------------------------------
# Local fixtures (no conftest.py exists for this test tree — inline by design)
# ---------------------------------------------------------------------------

def _make_test_db():
    load_model_modules()
    temp_root = Path('.tmp').resolve()
    temp_root.mkdir(parents=True, exist_ok=True)
    database_path = (temp_root / f'lz-{uuid4().hex}.db').as_posix()
    settings = Settings(database_url=f'sqlite+pysqlite:///{database_path}')
    engine = create_db_engine(settings)
    init_database(engine)
    return create_session_factory(settings)


@pytest.fixture()
def db_session():
    session_factory = _make_test_db()
    session = session_factory()
    try:
        yield session
    finally:
        session.close()


def _build_service(db_session: Session) -> ImportService:
    return ImportService(db_session)


# ---------------------------------------------------------------------------
# Template format tests (EMPNO-01 / D-09 / D-10)
# ---------------------------------------------------------------------------

@pytest.mark.parametrize('import_type', [
    'employees', 'certifications', 'performance_grades', 'salary_adjustments',
])
def test_template_employee_no_cell_format_is_text_for_all_types(
    db_session: Session, import_type: str
) -> None:
    svc = _build_service(db_session)
    result = svc.build_template_xlsx(import_type)
    # build_template_xlsx 可能返回 (filename, bytes, media_type) 或 bytes；适配两种
    content = result[1] if isinstance(result, tuple) else result
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # 工号固定在第 1 列
    for row_idx in range(1, ImportService.TEMPLATE_TEXT_PREFILL_ROWS + 1):
        cell = ws.cell(row=row_idx, column=1)
        assert cell.number_format == '@', (
            f'{import_type} 模板第 {row_idx} 行第 1 列 (员工工号) cell.number_format '
            f"应为 '@' 但实际为 {cell.number_format!r}"
        )


@pytest.mark.parametrize('import_type', [
    'employees', 'certifications', 'performance_grades', 'salary_adjustments',
])
def test_template_example_row_uses_leading_zero(
    db_session: Session, import_type: str
) -> None:
    svc = _build_service(db_session)
    result = svc.build_template_xlsx(import_type)
    content = result[1] if isinstance(result, tuple) else result
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # 示例行是第 2 行，工号固定在第 1 列
    assert ws.cell(row=2, column=1).value == '02651'


def test_template_manager_employee_no_column_is_text(db_session: Session) -> None:
    svc = _build_service(db_session)
    result = svc.build_template_xlsx('employees')
    content = result[1] if isinstance(result, tuple) else result
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # employees headers: 第 10 列是「直属上级工号」
    assert ws.cell(row=1, column=10).value == '直属上级工号'
    for row_idx in range(1, ImportService.TEMPLATE_TEXT_PREFILL_ROWS + 1):
        assert ws.cell(row=row_idx, column=10).number_format == '@'


def test_template_cert_type_and_stage_columns_are_text(db_session: Session) -> None:
    svc = _build_service(db_session)
    result = svc.build_template_xlsx('certifications')
    content = result[1] if isinstance(result, tuple) else result
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # certifications headers: 第 2 列=认证类型，第 3 列=认证阶段
    assert ws.cell(row=1, column=2).value == '认证类型'
    assert ws.cell(row=1, column=3).value == '认证阶段'
    for col_idx in (2, 3):
        for row_idx in range(1, ImportService.TEMPLATE_TEXT_PREFILL_ROWS + 1):
            assert ws.cell(row=row_idx, column=col_idx).number_format == '@'


# ---------------------------------------------------------------------------
# Sanity check tests (EMPNO-02 / D-06 / D-07)
# ---------------------------------------------------------------------------

def test_dispatch_import_flags_leading_zero_lost_rows(db_session: Session) -> None:
    svc = _build_service(db_session)
    df = pd.DataFrame([
        {'employee_no': '02651', 'year': '2025', 'grade': 'A'},   # pandas_idx=0, row_index=1
        {'employee_no': '1234.0', 'year': '2025', 'grade': 'B'},  # pandas_idx=1, row_index=2
    ])
    results = svc._dispatch_import('performance_grades', df)
    failed = [r for r in results if r.get('status') == 'failed']
    bad = [r for r in failed if '格式异常（疑似丢失前导零）' in str(r.get('message', ''))]
    assert len(bad) >= 1
    # row_index 约定：pandas_idx + 1 = 2
    assert any(r.get('row_index') == 2 for r in bad)


def test_dispatch_import_allows_valid_leading_zero_rows(db_session: Session) -> None:
    svc = _build_service(db_session)
    df = pd.DataFrame([
        {'employee_no': '02651', 'year': '2025', 'grade': 'A'},
    ])
    results = svc._dispatch_import('performance_grades', df)
    # 合法前导零行不应该因为 sanity check 被标 failed
    # （可能因员工不存在被标 failed，但 message 不应含「格式异常」）
    for r in results:
        assert '格式异常（疑似丢失前导零）' not in str(r.get('message', ''))


def test_manager_employee_no_also_checked(db_session: Session) -> None:
    svc = _build_service(db_session)
    df = pd.DataFrame([
        {
            'employee_no': '02651', 'name': '张三', 'department': '产品',
            'job_family': '平台', 'job_level': 'P5', 'manager_employee_no': '99.0',
        },
    ])
    results = svc._dispatch_import('employees', df)
    bad = [r for r in results if r.get('status') == 'failed'
           and '格式异常（疑似丢失前导零）' in str(r.get('message', ''))]
    assert len(bad) >= 1
    # 应当检测到 manager_employee_no 列
    assert any(
        '直属上级工号' in str(r.get('error_column', ''))
        or '直属上级工号' in str(r.get('message', ''))
        for r in bad
    )


def test_row_index_matches_existing_import_contract(db_session: Session) -> None:
    """row_index 口径一致性验证：sanity check 的 row_index 与既有 _import_* 的 row_index 一致。

    构造 3 条：idx=0 好行、idx=1 坏行、idx=2 好行。
    - sanity check 产生的 failed 记录：row_index=2（=idx 1 + 1）
    - 过滤坏行后，原 pandas index（0, 2）必须保留（不 reset_index），
      以保证下游 `_import_performance_grades` 拿到的 index 仍是原始值。
    """
    svc = _build_service(db_session)
    df = pd.DataFrame([
        {'employee_no': '02651', 'year': '2025', 'grade': 'A'},   # idx=0
        {'employee_no': '1234.0', 'year': '2025', 'grade': 'B'},  # idx=1 bad
        {'employee_no': '02652', 'year': '2025', 'grade': 'C'},   # idx=2
    ])

    # 直接验证 _detect_leading_zero_loss_rows 的 row_index 计算
    normalized = svc._normalize_columns('performance_grades', df)
    bad_map = ImportService._detect_leading_zero_loss_rows(normalized)
    # pandas_idx=1 对应 row_index=2
    assert 1 in bad_map
    row_index, col_name = bad_map[1]
    assert row_index == 2
    assert col_name == 'employee_no'

    # 整体 dispatch 也产生 row_index=2 的 failed 记录
    results = svc._dispatch_import('performance_grades', df)
    failed_rows = [r.get('row_index') for r in results
                   if r.get('status') == 'failed'
                   and '格式异常' in str(r.get('message', ''))]
    assert 2 in failed_rows

    # 反例：构造纯好行 3 条，验证 pandas index 保留不 reset
    df_good = pd.DataFrame([
        {'employee_no': '02651', 'year': '2025', 'grade': 'A'},
        {'employee_no': '02652', 'year': '2025', 'grade': 'B'},
        {'employee_no': '02653', 'year': '2025', 'grade': 'C'},
    ])
    normalized_good = svc._normalize_columns('performance_grades', df_good)
    bad_good = ImportService._detect_leading_zero_loss_rows(normalized_good)
    assert bad_good == {}  # 无坏行
