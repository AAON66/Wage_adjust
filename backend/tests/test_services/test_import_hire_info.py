"""Phase 32 Plan 02 Task 1: hire_info import 测试。

D-01 + D-02 + IMPORT-01 + IMPORT-05：
- SUPPORTED_TYPES 含 hire_info
- build_template_xlsx 可生成 hire_info 模板（工号列文本格式）
- _import_hire_info 支持 ISO 字符串日期 + Excel 序列号日期
- merge / replace 模式语义在必填字段上生效
"""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
import pytest
from openpyxl import load_workbook


pytestmark = pytest.mark.usefixtures('tmp_uploads_dir')


def test_supported_types_includes_hire_info():
    from backend.app.services.import_service import ImportService
    assert 'hire_info' in ImportService.SUPPORTED_TYPES
    assert 'non_statutory_leave' in ImportService.SUPPORTED_TYPES


def test_build_template_xlsx_hire_info(db_session):
    from backend.app.services.import_service import ImportService
    svc = ImportService(db_session)
    file_name, content, media_type = svc.build_template_xlsx('hire_info')
    assert content and len(content) > 0
    assert file_name.endswith('.xlsx')
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ['员工工号', '入职日期', '末次调薪日期']
    # 工号列文本格式（D-09）
    assert ws.cell(row=2, column=1).number_format == '@'


def test_import_hire_info_iso_date(db_session, employee_factory, xlsx_factory):
    from backend.app.services.import_service import ImportService
    emp = employee_factory(employee_no='E00001')
    rows = [['E00001', '2024-03-15', '2025-08-01']]
    data = xlsx_factory['hire_info'](rows=rows)

    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('hire_info', df)
    results = svc._import_hire_info(df)

    assert len(results) == 1
    assert results[0]['status'] == 'success'
    db_session.commit()
    db_session.refresh(emp)
    assert emp.hire_date == date(2024, 3, 15)
    assert emp.last_salary_adjustment_date == date(2025, 8, 1)


def test_import_hire_info_excel_serial_date(db_session, employee_factory, xlsx_factory):
    """Pitfall 1：Excel 序列号日期（dtype=str 后变成 '45292'）应正确解析。"""
    from backend.app.services.import_service import ImportService
    emp = employee_factory(employee_no='E00003')
    # 45292 = 2024-01-01 in Excel epoch (origin 1899-12-30)
    rows = [['E00003', 45292, '2025-06-01']]
    data = xlsx_factory['hire_info'](rows=rows)

    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('hire_info', df)
    results = svc._import_hire_info(df)

    assert results[0]['status'] == 'success', results
    db_session.commit()
    db_session.refresh(emp)
    assert emp.hire_date == date(2024, 1, 1)


def test_import_hire_info_employee_not_found(db_session, xlsx_factory):
    from backend.app.services.import_service import ImportService
    rows = [['E99999', '2024-03-15', None]]
    data = xlsx_factory['hire_info'](rows=rows)

    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('hire_info', df)
    results = svc._import_hire_info(df)

    assert results[0]['status'] == 'failed'
    assert '未找到员工工号' in results[0]['error']


def test_import_hire_info_all_empty_no_change(db_session, employee_factory, xlsx_factory):
    """hire_date 与 last_salary_adjustment_date 都为空 → no_change（merge 模式）。"""
    from backend.app.services.import_service import ImportService
    emp = employee_factory(employee_no='E00001', hire_date=date(2020, 1, 1))
    rows = [['E00001', None, None]]
    data = xlsx_factory['hire_info'](rows=rows)

    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('hire_info', df)
    results = svc._import_hire_info(df, overwrite_mode='merge')

    assert results[0]['status'] == 'success'
    assert results[0]['action'] == 'no_change'
    db_session.refresh(emp)
    assert emp.hire_date == date(2020, 1, 1)


def test_import_hire_info_replace_mode_required_field_empty(
    db_session, employee_factory, xlsx_factory
):
    """D-12: replace 模式下必填 hire_date 空 → row failed。"""
    from backend.app.services.import_service import ImportService
    employee_factory(employee_no='E00001', hire_date=date(2020, 1, 1))
    rows = [['E00001', None, None]]
    data = xlsx_factory['hire_info'](rows=rows)

    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('hire_info', df)
    results = svc._import_hire_info(df, overwrite_mode='replace')

    assert results[0]['status'] == 'failed'
    assert '入职日期' in results[0]['error']
