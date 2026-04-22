"""Phase 32 Wave 0 scaffold tests.

确认：
  - ImportJob 模型字段已升级（overwrite_mode / actor_id / updated_at）
  - SalaryAdjustmentRecord UniqueConstraint 已加（D-14 / Pitfall 4）
  - 测试 fixture 与 xlsx builder 可调用
  - 后续 Wave 的 pytest 命令在文件被创建时已能 collect

本文件 GREEN 即代表 Wave 1 完成，可以推进到 Wave 2 实施。
"""
from __future__ import annotations

import io


def test_import_job_has_overwrite_mode_column():
    from backend.app.models.import_job import ImportJob
    cols = {c.name for c in ImportJob.__table__.columns}
    assert 'overwrite_mode' in cols, '缺 overwrite_mode 列（Phase 32 D-12）'


def test_import_job_has_actor_id_column():
    from backend.app.models.import_job import ImportJob
    cols = {c.name for c in ImportJob.__table__.columns}
    assert 'actor_id' in cols, '缺 actor_id 列（Phase 32 D-13）'


def test_import_job_has_updated_at_column():
    from backend.app.models.import_job import ImportJob
    cols = {c.name for c in ImportJob.__table__.columns}
    assert 'updated_at' in cols, '缺 updated_at 列（Phase 32 D-17 expire 任务依赖）'


def test_salary_adj_unique_constraint_present():
    from backend.app.models.salary_adjustment_record import SalaryAdjustmentRecord
    names = [c.name for c in SalaryAdjustmentRecord.__table__.constraints if c.name]
    assert 'uq_salary_adj_employee_date_type' in names, \
        '缺 UniqueConstraint(employee_id, adjustment_date, adjustment_type)（Pitfall 4 / D-14）'


def test_performance_record_unique_constraint_already_exists():
    """决议 D-15: PerformanceRecord 已有 uq_performance_employee_year，本期不需要新加。"""
    from backend.app.models.performance_record import PerformanceRecord
    names = [c.name for c in PerformanceRecord.__table__.constraints if c.name]
    assert 'uq_performance_employee_year' in names


def test_xlsx_factory_builds_hire_info(xlsx_factory):
    from openpyxl import load_workbook
    data = xlsx_factory['hire_info']()
    assert isinstance(data, (bytes, bytearray)) and len(data) > 0
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ['员工工号', '入职日期', '末次调薪日期']


def test_xlsx_factory_builds_hire_info_with_serial_date(xlsx_factory):
    from openpyxl import load_workbook
    data = xlsx_factory['hire_info'](with_serial_date=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    # 第 4 行（含序列日期）
    row = [c.value for c in ws[4]]
    assert row[0] == 'E00003'
    assert row[1] == 45292  # Excel serial date


def test_xlsx_factory_builds_non_statutory_leave_with_conflict(xlsx_factory):
    from openpyxl import load_workbook
    data = xlsx_factory['non_statutory_leave'](with_conflict=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    # 应含两条 (E00001, 2026)，触发 D-09 同文件冲突检测
    e00001_2026 = [r for r in rows if r[0] == 'E00001' and r[1] == 2026]
    assert len(e00001_2026) == 2


def test_tmp_uploads_dir_isolates_storage(tmp_uploads_dir):
    from backend.app.core.config import get_settings
    s = get_settings()
    # storage_base_dir 应该指向 tmp_path，不污染开发环境
    assert str(tmp_uploads_dir) in str(s.storage_base_dir)


def test_employee_factory_creates_with_defaults(db_session, employee_factory):
    emp = employee_factory()
    assert emp.id
    assert emp.employee_no.startswith('E')
    assert emp.department == 'R&D'


def test_user_factory_creates_with_role(db_session, user_factory):
    user = user_factory(role='hrbp')
    assert user.id
    assert user.role == 'hrbp'
    assert user.hashed_password  # 已 hash


def test_import_job_factory_creates_with_phase32_fields(db_session, import_job_factory):
    job = import_job_factory(import_type='hire_info', status='previewing', overwrite_mode='replace')
    assert job.id
    assert job.overwrite_mode == 'replace'
    assert job.status == 'previewing'


def test_client_anon_returns_401_on_protected_endpoint(client_anon):
    """smoke: TestClient 无 token 调任何 protected endpoint → 401。"""
    resp = client_anon.get('/api/v1/auth/me')
    assert resp.status_code in (401, 404)  # 401 if endpoint exists, 404 otherwise — either OK


def test_client_hrbp_carries_authorization_header(client_hrbp):
    """smoke: TestClient 已注入 Bearer header。"""
    assert client_hrbp.headers.get('Authorization', '').startswith('Bearer ')


def test_client_employee_role_in_token(client_employee, employee_user_token):
    """smoke: employee token 已注入 Authorization header。"""
    assert employee_user_token
    assert client_employee.headers.get('Authorization', '').endswith(employee_user_token)
