"""Phase 32 Plan 02 Task 1: non_statutory_leave import 测试。

D-03 + IMPORT-01 + IMPORT-05：
- _import_non_statutory_leave upsert by (employee_id, year)
- Decimal 精度（避免 float 误差）
- leave_type 可选
"""
from __future__ import annotations

import io
from decimal import Decimal

import pandas as pd
import pytest
from sqlalchemy import select


pytestmark = pytest.mark.usefixtures('tmp_uploads_dir')


def test_build_template_xlsx_non_statutory_leave(db_session):
    from openpyxl import load_workbook
    from backend.app.services.import_service import ImportService
    svc = ImportService(db_session)
    file_name, content, _ = svc.build_template_xlsx('non_statutory_leave')
    assert content and len(content) > 0
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == ['员工工号', '年度', '假期天数', '假期类型']
    assert ws.cell(row=2, column=1).number_format == '@'


def test_import_non_statutory_leave_upsert_by_year(
    db_session, employee_factory, xlsx_factory
):
    """业务键 (employee_id, year)：相同年度第二次导入 → update 而非 insert。"""
    from backend.app.services.import_service import ImportService
    from backend.app.models.non_statutory_leave import NonStatutoryLeave

    emp = employee_factory(employee_no='E00001')
    svc = ImportService(db_session)

    # 第一次：insert
    data = xlsx_factory['non_statutory_leave'](rows=[['E00001', 2026, 10.5, '事假']])
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('non_statutory_leave', df)
    r1 = svc._import_non_statutory_leave(df)
    db_session.commit()
    assert r1[0]['status'] == 'success'
    assert r1[0]['action'] == 'insert'

    # 第二次：相同 (employee_id, year) → update
    data2 = xlsx_factory['non_statutory_leave'](rows=[['E00001', 2026, 12.0, '病假']])
    df2 = pd.read_excel(io.BytesIO(data2), dtype=str).fillna('')
    df2 = svc._normalize_columns('non_statutory_leave', df2)
    r2 = svc._import_non_statutory_leave(df2)
    db_session.commit()
    assert r2[0]['action'] == 'update'

    # DB 中只有一条记录
    records = db_session.execute(
        select(NonStatutoryLeave).where(
            NonStatutoryLeave.employee_id == emp.id,
            NonStatutoryLeave.year == 2026,
        )
    ).scalars().all()
    assert len(records) == 1
    assert records[0].total_days == Decimal('12.0')
    assert records[0].leave_type == '病假'


def test_import_non_statutory_leave_decimal_precision(
    db_session, employee_factory, xlsx_factory
):
    """Pitfall 8：total_days 用 Decimal 持久化，无 float 误差。"""
    from backend.app.services.import_service import ImportService
    from backend.app.models.non_statutory_leave import NonStatutoryLeave

    emp = employee_factory(employee_no='E00002')
    svc = ImportService(db_session)
    data = xlsx_factory['non_statutory_leave'](rows=[['E00002', 2026, '0.5', None]])
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('non_statutory_leave', df)
    svc._import_non_statutory_leave(df)
    db_session.commit()

    rec = db_session.execute(
        select(NonStatutoryLeave).where(NonStatutoryLeave.employee_id == emp.id)
    ).scalar_one()
    assert rec.total_days == Decimal('0.5')


def test_import_non_statutory_leave_optional_leave_type(
    db_session, employee_factory, xlsx_factory
):
    """leave_type 为空 → 仍可写入（可选字段）。"""
    from backend.app.services.import_service import ImportService
    from backend.app.models.non_statutory_leave import NonStatutoryLeave

    emp = employee_factory(employee_no='E00003')
    svc = ImportService(db_session)
    data = xlsx_factory['non_statutory_leave'](rows=[['E00003', 2026, 5.0, None]])
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('non_statutory_leave', df)
    results = svc._import_non_statutory_leave(df)
    db_session.commit()

    assert results[0]['status'] == 'success'
    rec = db_session.execute(
        select(NonStatutoryLeave).where(NonStatutoryLeave.employee_id == emp.id)
    ).scalar_one()
    assert rec.leave_type is None
    assert rec.total_days == Decimal('5.0')


def test_import_non_statutory_leave_employee_not_found(db_session, xlsx_factory):
    from backend.app.services.import_service import ImportService
    data = xlsx_factory['non_statutory_leave'](rows=[['E99999', 2026, 10.0, '事假']])
    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('non_statutory_leave', df)
    results = svc._import_non_statutory_leave(df)

    assert results[0]['status'] == 'failed'
    assert '未找到员工工号' in results[0]['error']


def test_import_non_statutory_leave_invalid_year(
    db_session, employee_factory, xlsx_factory
):
    from backend.app.services.import_service import ImportService
    employee_factory(employee_no='E00004')
    data = xlsx_factory['non_statutory_leave'](rows=[['E00004', 'abc', 5.0, None]])
    svc = ImportService(db_session)
    df = pd.read_excel(io.BytesIO(data), dtype=str).fillna('')
    df = svc._normalize_columns('non_statutory_leave', df)
    results = svc._import_non_statutory_leave(df)
    assert results[0]['status'] == 'failed'
    assert '年度' in results[0]['error']
