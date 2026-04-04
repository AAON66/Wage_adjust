"""Tests for EligibilityService batch query with filter-before-paginate."""
from __future__ import annotations

from datetime import date, datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from backend.app.core.database import Base
from backend.app.models import load_model_modules
from backend.app.models.employee import Employee
from backend.app.models.attendance_record import AttendanceRecord
from backend.app.models.performance_record import PerformanceRecord
from backend.app.models.salary_adjustment_record import SalaryAdjustmentRecord
from backend.app.models.user import User
from backend.app.models.department import Department
from backend.app.services.eligibility_service import EligibilityService


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _make_db() -> tuple[Session, sessionmaker]:
    temp_root = Path('.tmp').resolve()
    temp_root.mkdir(parents=True, exist_ok=True)
    db_path = (temp_root / f'elig-batch-{uuid4().hex}.db').as_posix()
    engine = create_engine(
        f'sqlite+pysqlite:///{db_path}',
        connect_args={'check_same_thread': False},
        echo=False,
    )
    load_model_modules()
    Base.metadata.create_all(bind=engine)
    factory = sessionmaker(bind=engine, autocommit=False, autoflush=False)
    return factory(), factory


def _make_employee(db: Session, *, employee_no: str, name: str, department: str,
                   job_family: str = 'tech', job_level: str = 'P5',
                   hire_date: date | None = date(2020, 1, 1)) -> Employee:
    emp = Employee(
        id=uuid4().hex[:36],
        employee_no=employee_no,
        name=name,
        department=department,
        job_family=job_family,
        job_level=job_level,
        hire_date=hire_date,
    )
    db.add(emp)
    db.flush()
    return emp


def _make_user(db: Session, *, role: str, department_names: list[str] | None = None,
               employee_id: str | None = None) -> User:
    user = User(
        id=uuid4().hex[:36],
        email=f'{uuid4().hex[:8]}@test.com',
        hashed_password='x',
        role=role,
        employee_id=employee_id,
    )
    db.add(user)
    db.flush()
    if department_names:
        for dept_name in department_names:
            dept = db.query(Department).filter(Department.name == dept_name).first()
            if dept is None:
                dept = Department(id=uuid4().hex[:36], name=dept_name)
                db.add(dept)
                db.flush()
            user.departments.append(dept)
        db.flush()
    return user


def _make_perf(db: Session, emp: Employee, year: int, grade: str) -> None:
    rec = PerformanceRecord(
        employee_id=emp.id,
        employee_no=emp.employee_no,
        year=year,
        grade=grade,
    )
    db.add(rec)
    db.flush()


def _make_attendance(db: Session, emp: Employee, period: str, leave_days: float) -> None:
    now = datetime.now(tz=timezone.utc)
    rec = AttendanceRecord(
        employee_id=emp.id,
        employee_no=emp.employee_no,
        period=period,
        non_statutory_leave_days=leave_days,
        data_as_of=now,
        synced_at=now,
    )
    db.add(rec)
    db.flush()


def _make_salary_adj(db: Session, emp: Employee, adj_date: date) -> None:
    rec = SalaryAdjustmentRecord(
        employee_id=emp.id,
        employee_no=emp.employee_no,
        adjustment_date=adj_date,
        adjustment_type='annual',
    )
    db.add(rec)
    db.flush()


# ---------------------------------------------------------------------------
# Tests: batch query
# ---------------------------------------------------------------------------

class TestCheckEmployeesBatch:
    def setup_method(self) -> None:
        self.db, self.factory = _make_db()
        self.ref_date = date(2026, 4, 1)
        self.year = 2025

        # Create admin user
        self.admin = _make_user(self.db, role='admin')

        # Create employees:
        # E1: eligible (good tenure, good perf, no excess leave)
        self.e1 = _make_employee(self.db, employee_no='E001', name='Alice', department='Engineering')
        _make_perf(self.db, self.e1, self.year, 'A')
        _make_attendance(self.db, self.e1, f'{self.year}-Q4', 5.0)

        # E2: ineligible (bad perf)
        self.e2 = _make_employee(self.db, employee_no='E002', name='Bob', department='Engineering')
        _make_perf(self.db, self.e2, self.year, 'D')
        _make_attendance(self.db, self.e2, f'{self.year}-Q4', 5.0)

        # E3: ineligible (excess leave)
        self.e3 = _make_employee(self.db, employee_no='E003', name='Carol', department='Engineering')
        _make_perf(self.db, self.e3, self.year, 'A')
        _make_attendance(self.db, self.e3, f'{self.year}-Q4', 40.0)

        # E4: eligible (different department)
        self.e4 = _make_employee(self.db, employee_no='E004', name='Dave', department='Sales')
        _make_perf(self.db, self.e4, self.year, 'A')
        _make_attendance(self.db, self.e4, f'{self.year}-Q4', 5.0)

        # E5: ineligible (bad perf, Sales)
        self.e5 = _make_employee(self.db, employee_no='E005', name='Eve', department='Sales')
        _make_perf(self.db, self.e5, self.year, 'D')
        _make_attendance(self.db, self.e5, f'{self.year}-Q4', 5.0)

        self.db.commit()

    def teardown_method(self) -> None:
        self.db.close()

    def _service(self) -> EligibilityService:
        return EligibilityService(self.db)

    def test_batch_returns_dict_for_multiple_employees(self) -> None:
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter=None, rule_filter=None,
            job_family=None, job_level=None,
            page=1, page_size=50,
            current_user=self.admin,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 5
        assert len(items) == 5

    def test_filter_before_paginate_status_filter(self) -> None:
        """HIGH concern #2: filter-then-paginate. status_filter='ineligible', page_size=2.
        We have 3 ineligible employees (E2, E3, E5). total should be 3, not 5."""
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter='ineligible', rule_filter=None,
            job_family=None, job_level=None,
            page=1, page_size=2,
            current_user=self.admin,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 3, f'Expected 3 ineligible employees, got total={total}'
        assert len(items) == 2, 'Page size is 2, should return 2 items'

    def test_filter_before_paginate_page_2(self) -> None:
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter='ineligible', rule_filter=None,
            job_family=None, job_level=None,
            page=2, page_size=2,
            current_user=self.admin,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 3
        assert len(items) == 1, 'Page 2 of 3 items with page_size=2 should return 1'

    def test_rule_filter(self) -> None:
        """rule_filter='PERFORMANCE' should only return employees where PERFORMANCE is ineligible."""
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter=None, rule_filter='PERFORMANCE',
            job_family=None, job_level=None,
            page=1, page_size=50,
            current_user=self.admin,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 2  # E2 and E5 have bad performance
        for item in items:
            perf_rules = [r for r in item['rules'] if r['rule_code'] == 'PERFORMANCE']
            assert perf_rules[0]['status'] == 'ineligible'

    def test_department_filter(self) -> None:
        svc = self._service()
        items, total = svc.check_employees_batch(
            department='Sales', status_filter=None, rule_filter=None,
            job_family=None, job_level=None,
            page=1, page_size=50,
            current_user=self.admin,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 2  # E4, E5

    def test_manager_scope_department(self) -> None:
        """Manager should only see employees in their department via AccessScopeService."""
        manager = _make_user(self.db, role='manager', department_names=['Engineering'])
        self.db.commit()
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter=None, rule_filter=None,
            job_family=None, job_level=None,
            page=1, page_size=50,
            current_user=manager,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 3  # E1, E2, E3 (Engineering only)

    def test_hrbp_scope_department(self) -> None:
        """HRBP scoped same as manager (review MEDIUM #6)."""
        hrbp = _make_user(self.db, role='hrbp', department_names=['Sales'])
        self.db.commit()
        svc = self._service()
        items, total = svc.check_employees_batch(
            department=None, status_filter=None, rule_filter=None,
            job_family=None, job_level=None,
            page=1, page_size=50,
            current_user=hrbp,
            reference_date=self.ref_date, year=self.year,
        )
        assert total == 2  # E4, E5 (Sales only)


class TestExportEligibilityExcel:
    def setup_method(self) -> None:
        self.db, _ = _make_db()

    def teardown_method(self) -> None:
        self.db.close()

    def test_export_produces_valid_xlsx(self) -> None:
        items = [
            {
                'employee_no': 'E001',
                'name': 'Alice',
                'department': 'Engineering',
                'job_family': 'tech',
                'job_level': 'P5',
                'overall_status': 'eligible',
                'rules': [
                    {'rule_code': 'TENURE', 'rule_label': '入职时长', 'status': 'eligible', 'detail': '已入职 72 个月'},
                    {'rule_code': 'ADJUSTMENT_INTERVAL', 'rule_label': '调薪间隔', 'status': 'data_missing', 'detail': '无调薪记录'},
                    {'rule_code': 'PERFORMANCE', 'rule_label': '绩效等级', 'status': 'eligible', 'detail': '绩效等级为 A'},
                    {'rule_code': 'LEAVE', 'rule_label': '非法定假期', 'status': 'eligible', 'detail': '非法定假期 5.0 天'},
                ],
            },
        ]
        svc = EligibilityService(self.db)
        buf = svc.export_eligibility_excel(items)
        assert buf is not None
        # Verify it's valid xlsx
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert '工号' in headers
        assert '姓名' in headers
        assert '资格状态' in headers

    def test_export_max_5000_rows(self) -> None:
        items = [
            {
                'employee_no': f'E{i:04d}',
                'name': f'Name{i}',
                'department': 'Eng',
                'job_family': 'tech',
                'job_level': 'P5',
                'overall_status': 'eligible',
                'rules': [],
            }
            for i in range(6000)
        ]
        svc = EligibilityService(self.db)
        buf = svc.export_eligibility_excel(items)
        import openpyxl
        wb = openpyxl.load_workbook(buf)
        ws = wb.active
        # 1 header + max 5000 data rows
        assert ws.max_row <= 5001
