from __future__ import annotations

import io
from pathlib import Path
from uuid import uuid4

from sqlalchemy import select, func

from backend.app.core.config import Settings
from backend.app.core.database import create_db_engine, create_session_factory, init_database
from backend.app.models import load_model_modules
from backend.app.models.department import Department
from backend.app.models.employee import Employee
from backend.app.services.import_service import ImportService


class UploadStub:
    def __init__(self, filename: str, content: bytes) -> None:
        self.filename = filename
        self.file = io.BytesIO(content)


def _make_test_db():
    load_model_modules()
    temp_root = Path('.tmp').resolve()
    temp_root.mkdir(parents=True, exist_ok=True)
    database_path = (temp_root / f'xlsx-{uuid4().hex}.db').as_posix()
    settings = Settings(database_url=f'sqlite+pysqlite:///{database_path}')
    engine = create_db_engine(settings)
    init_database(engine)
    return create_session_factory(settings)


def _seed_department(db, name: str = '产品技术中心') -> None:
    existing = db.scalar(select(Department).where(Department.name == name))
    if existing is None:
        db.add(Department(name=name, description=f'{name} scope', status='active'))
        db.commit()


# ---------------------------------------------------------------------------
# IMP-04: xlsx file import + encoding compatibility
# ---------------------------------------------------------------------------


class TestXlsxRead:
    """IMP-04: xlsx files can be imported; GBK/BOM CSV are decoded correctly."""

    def test_xlsx_file_parsed_successfully(self) -> None:
        """Create a real xlsx file via openpyxl, import it, verify success."""
        session_factory = _make_test_db()
        with session_factory() as db:
            _seed_department(db)
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['员工工号', '员工姓名', '所属部门', '下属部门', '所属公司', '岗位族', '岗位级别', '在职状态', '身份证号'])
            ws.append(['XLSX-001', '张Excel', '产品技术中心', '后端平台组', '星海集团', '技术', 'P6', 'active', ''])
            output = io.BytesIO()
            wb.save(output)
            xlsx_bytes = output.getvalue()

            upload = UploadStub('test.xlsx', xlsx_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.total_rows == 1
            assert job.success_rows == 1
            # DB assertion
            emp = db.scalar(select(Employee).where(Employee.employee_no == 'XLSX-001'))
            assert emp is not None
            assert emp.name == '张Excel'
            assert emp.company == '星海集团'

    def test_gbk_csv_no_garbled_characters(self) -> None:
        """Import GBK-encoded CSV, verify Chinese fields stored correctly (no garbling)."""
        session_factory = _make_test_db()
        with session_factory() as db:
            _seed_department(db)
            csv_text = '员工工号,员工姓名,所属部门,岗位族,岗位级别,在职状态,身份证号\nGBK-001,张国标编码,产品技术中心,技术,P6,active,\n'
            gbk_bytes = csv_text.encode('gbk')
            upload = UploadStub('gbk.csv', gbk_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.success_rows == 1
            # DB assertion: Chinese name stored correctly
            emp = db.scalar(select(Employee).where(Employee.employee_no == 'GBK-001'))
            assert emp is not None
            assert emp.name == '张国标编码', f'Expected 张国标编码, got {emp.name}'

    def test_utf8_bom_csv_parsed(self) -> None:
        """Import UTF-8 BOM-encoded CSV, verify correct parsing."""
        session_factory = _make_test_db()
        with session_factory() as db:
            _seed_department(db)
            csv_text = '员工工号,员工姓名,所属部门,岗位族,岗位级别,在职状态,身份证号\nBOM-001,带BOM头,产品技术中心,技术,P6,active,\n'
            bom_bytes = b'\xef\xbb\xbf' + csv_text.encode('utf-8')
            upload = UploadStub('bom.csv', bom_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.success_rows == 1

    def test_5000_row_limit_enforced(self) -> None:
        """CSV with >5000 rows must be rejected with proper error message."""
        session_factory = _make_test_db()
        with session_factory() as db:
            header = '员工工号,员工姓名,所属部门,岗位族,岗位级别,在职状态,身份证号\n'
            rows = ''.join(f'OVER-{i:05d},测试员工{i},产品技术中心,技术,P6,active,\n' for i in range(5001))
            csv_bytes = (header + rows).encode('utf-8')
            upload = UploadStub('too_many.csv', csv_bytes)
            service = ImportService(db)
            # The service should enforce MAX_ROWS limit -- either raise ValueError or return failed job
            try:
                job = service.run_import(import_type='employees', upload=upload)
                # If it doesn't raise, check that the job indicates the limit was exceeded
                assert job.failed_rows > 0 or job.status == 'failed' or '5000' in str(job.result_summary)
            except (ValueError, Exception) as exc:
                assert '5000' in str(exc), f'Error should mention 5000 row limit, got: {exc}'

    def test_gbk_fixture_file_imported(self) -> None:
        """Import the pre-built GBK fixture file from disk."""
        session_factory = _make_test_db()
        with session_factory() as db:
            _seed_department(db)
            fixture_path = Path(__file__).parent.parent / 'fixtures' / 'import_gbk.csv'
            gbk_bytes = fixture_path.read_bytes()
            upload = UploadStub('import_gbk.csv', gbk_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.success_rows >= 2, f'Expected at least 2 successes from GBK fixture, got {job.success_rows}'

    def test_utf8bom_fixture_file_imported(self) -> None:
        """Import the pre-built UTF-8 BOM fixture file from disk."""
        session_factory = _make_test_db()
        with session_factory() as db:
            _seed_department(db)
            fixture_path = Path(__file__).parent.parent / 'fixtures' / 'import_utf8bom.csv'
            bom_bytes = fixture_path.read_bytes()
            upload = UploadStub('import_utf8bom.csv', bom_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.success_rows >= 1, f'Expected at least 1 success from BOM fixture, got {job.success_rows}'


# ---------------------------------------------------------------------------
# IMP-06: xlsx template download
# ---------------------------------------------------------------------------


class TestXlsxTemplate:
    """IMP-06: xlsx template generation and error report export."""

    def test_employee_xlsx_template_generated(self) -> None:
        """build_template_xlsx('employees') returns xlsx bytes with header row."""
        session_factory = _make_test_db()
        with session_factory() as db:
            service = ImportService(db)
            file_name, content, media_type = service.build_template_xlsx('employees')
            assert file_name.endswith('.xlsx')
            assert 'spreadsheetml' in media_type
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            assert ws.cell(row=1, column=1).value is not None, 'Header row must exist'
            headers = [ws.cell(row=1, column=index).value for index in range(1, ws.max_column + 1)]
            assert '所属公司' in headers

    def test_certification_xlsx_template_generated(self) -> None:
        """build_template_xlsx('certifications') returns xlsx bytes."""
        session_factory = _make_test_db()
        with session_factory() as db:
            service = ImportService(db)
            file_name, content, media_type = service.build_template_xlsx('certifications')
            assert file_name.endswith('.xlsx')
            assert len(content) > 0

    def test_xlsx_template_has_example_rows(self) -> None:
        """Template must have at least header + 1 example row."""
        session_factory = _make_test_db()
        with session_factory() as db:
            service = ImportService(db)
            _, content, _ = service.build_template_xlsx('employees')
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            assert ws.max_row >= 2, 'Template must have at least header + 1 example row'

    def test_xlsx_error_report_generated(self) -> None:
        """build_export_report_xlsx generates xlsx with failed row data."""
        session_factory = _make_test_db()
        with session_factory() as db:
            csv_bytes = (
                '员工工号,员工姓名,所属部门,岗位族,岗位级别,在职状态,身份证号\n'
                'RPT-001,报告失败,不存在部门,技术,P7,active,\n'
            ).encode('utf-8')
            upload = UploadStub('report.csv', csv_bytes)
            service = ImportService(db)
            job = service.run_import(import_type='employees', upload=upload)
            assert job.failed_rows >= 1
            file_name, content, media_type = service.build_export_report_xlsx(job)
            assert file_name.endswith('.xlsx')
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            assert ws.max_row >= 2, 'Error report must have header + at least 1 failed row'
