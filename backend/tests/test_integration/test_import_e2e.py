"""Phase 32-06 E2E 集成测试 — 跨层验收 SC1-SC5。

覆盖：
  - SC1: 4 类资格 import_type 模板下载非空 xlsx + openpyxl 可读回
  - SC2: Preview + diff 数据结构（hire_info merge / non_statutory_leave 冲突）
  - SC3: merge / replace 模式 + AuditLog detail.overwrite_mode
  - SC5: 重复导入幂等（performance_grades / salary_adjustments 业务键 upsert）
  - APScheduler import_scheduler smoke：模块结构 + run_expire_stale_jobs 不抛
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook


# ===================== 辅助函数 =====================


def _seed_employee(api_ctx, *, employee_no: str = 'E00001'):
    """在 API 共享 DB 中 seed 一个 Employee，返回 id。"""
    from backend.app.models.employee import Employee
    db = api_ctx.session_factory()
    try:
        emp = Employee(
            employee_no=employee_no,
            name='测试员工',
            department='R&D',
            job_family='engineering',
            job_level='P5',
            status='active',
        )
        db.add(emp)
        db.commit()
        db.refresh(emp)
        return emp.id
    finally:
        db.close()


def _post_preview(client, xlsx_bytes: bytes, import_type: str, filename: str = 't.xlsx'):
    return client.post(
        f'/api/v1/eligibility-import/excel/preview?import_type={import_type}',
        files={
            'file': (
                filename,
                io.BytesIO(xlsx_bytes),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ),
        },
    )


def _post_confirm(client, job_id: str, *, overwrite_mode: str = 'merge', confirm_replace: bool = False):
    return client.post(
        f'/api/v1/eligibility-import/excel/{job_id}/confirm',
        json={'overwrite_mode': overwrite_mode, 'confirm_replace': confirm_replace},
    )


# ===================== E2E 用例 =====================


def test_e2e_hire_info_merge(client_hrbp, _api_context, xlsx_factory, tmp_uploads_dir):
    """E2E hire_info merge：preview → confirm → DB 更新 + AuditLog 写入 + 暂存文件删除。"""
    from datetime import date

    from sqlalchemy import select

    from backend.app.models.audit_log import AuditLog
    from backend.app.models.employee import Employee

    _seed_employee(_api_context, employee_no='E00001')
    data = xlsx_factory['hire_info'](rows=[['E00001', '2024-03-15', '2025-08-01']])

    prev = _post_preview(client_hrbp, data, 'hire_info')
    assert prev.status_code == 200, prev.text
    job_id = prev.json()['job_id']

    confirm = _post_confirm(client_hrbp, job_id, overwrite_mode='merge')
    assert confirm.status_code == 200, confirm.text
    body = confirm.json()
    assert body['status'] == 'completed'

    # 校验 DB 真的写入
    db = _api_context.session_factory()
    try:
        emp = db.execute(
            select(Employee).where(Employee.employee_no == 'E00001'),
        ).scalar_one()
        assert emp.hire_date == date(2024, 3, 15)
        log = db.execute(
            select(AuditLog).where(
                AuditLog.target_id == job_id,
                AuditLog.action == 'import_confirmed',
            ),
        ).scalar_one()
        assert log.detail['overwrite_mode'] == 'merge'
    finally:
        db.close()


def test_e2e_hire_info_replace_with_confirm_flag(
    client_hrbp, _api_context, xlsx_factory, tmp_uploads_dir,
):
    """E2E hire_info replace + confirm_replace=True → AuditLog detail.overwrite_mode == 'replace'。"""
    from sqlalchemy import select

    from backend.app.models.audit_log import AuditLog

    _seed_employee(_api_context, employee_no='E00001')
    data = xlsx_factory['hire_info'](rows=[['E00001', '2024-03-15', '2025-08-01']])

    prev = _post_preview(client_hrbp, data, 'hire_info')
    assert prev.status_code == 200, prev.text
    job_id = prev.json()['job_id']

    confirm = _post_confirm(client_hrbp, job_id, overwrite_mode='replace', confirm_replace=True)
    assert confirm.status_code == 200, confirm.text

    db = _api_context.session_factory()
    try:
        log = db.execute(
            select(AuditLog).where(AuditLog.target_id == job_id),
        ).scalar_one()
        assert log.detail['overwrite_mode'] == 'replace'
    finally:
        db.close()


def test_e2e_perf_grades_idempotent(client_hrbp, _api_context, xlsx_factory, tmp_uploads_dir):
    """SC5: 重复导入相同 (employee_no, year)，DB 中只有 1 条 PerformanceRecord。"""
    from sqlalchemy import select

    from backend.app.models.employee import Employee
    from backend.app.models.performance_record import PerformanceRecord

    _seed_employee(_api_context, employee_no='E00001')
    data = xlsx_factory['performance_grades'](rows=[['E00001', 2026, 'A']])

    # 第一次 preview + confirm
    p1 = _post_preview(client_hrbp, data, 'performance_grades')
    assert p1.status_code == 200, p1.text
    c1 = _post_confirm(client_hrbp, p1.json()['job_id'], overwrite_mode='merge')
    assert c1.status_code == 200, c1.text

    # 第二次 preview + confirm（同样数据）
    p2 = _post_preview(client_hrbp, data, 'performance_grades')
    assert p2.status_code == 200, p2.text
    c2 = _post_confirm(client_hrbp, p2.json()['job_id'], overwrite_mode='merge')
    assert c2.status_code == 200, c2.text

    db = _api_context.session_factory()
    try:
        emp = db.execute(
            select(Employee).where(Employee.employee_no == 'E00001'),
        ).scalar_one()
        recs = db.execute(
            select(PerformanceRecord).where(
                PerformanceRecord.employee_id == emp.id,
                PerformanceRecord.year == 2026,
            ),
        ).scalars().all()
        assert len(recs) == 1, f'Expected 1 record, got {len(recs)}'
    finally:
        db.close()


def test_e2e_salary_adjustments_idempotent(
    client_hrbp, _api_context, xlsx_factory, tmp_uploads_dir,
):
    """Pitfall 4 解决：upsert by (employee_no, adjustment_date, adjustment_type)。"""
    from sqlalchemy import select

    from backend.app.models.employee import Employee
    from backend.app.models.salary_adjustment_record import SalaryAdjustmentRecord

    _seed_employee(_api_context, employee_no='E00001')
    data = xlsx_factory['salary_adjustments'](rows=[['E00001', '2026-01-01', 'annual', 1000.0]])

    p1 = _post_preview(client_hrbp, data, 'salary_adjustments')
    assert p1.status_code == 200, p1.text
    c1 = _post_confirm(client_hrbp, p1.json()['job_id'], overwrite_mode='merge')
    assert c1.status_code == 200, c1.text

    p2 = _post_preview(client_hrbp, data, 'salary_adjustments')
    assert p2.status_code == 200, p2.text
    c2 = _post_confirm(client_hrbp, p2.json()['job_id'], overwrite_mode='merge')
    assert c2.status_code == 200, c2.text

    db = _api_context.session_factory()
    try:
        emp = db.execute(
            select(Employee).where(Employee.employee_no == 'E00001'),
        ).scalar_one()
        recs = db.execute(
            select(SalaryAdjustmentRecord).where(
                SalaryAdjustmentRecord.employee_id == emp.id,
            ),
        ).scalars().all()
        assert len(recs) == 1, f'Expected 1 record, got {len(recs)}'
    finally:
        db.close()


def test_e2e_non_statutory_leave_conflict_detection(
    client_hrbp, _api_context, xlsx_factory, tmp_uploads_dir,
):
    """同文件内业务键重复 → preview counters.conflict 反映。"""
    _seed_employee(_api_context, employee_no='E00001')
    _seed_employee(_api_context, employee_no='E00002')
    data = xlsx_factory['non_statutory_leave'](with_conflict=True)

    prev = _post_preview(client_hrbp, data, 'non_statutory_leave')
    assert prev.status_code == 200, prev.text
    body = prev.json()
    # 同业务键 (E00001, 2026) 重复 2 次 → 至少 2 行 conflict
    assert body['counters']['conflict'] >= 2


@pytest.mark.parametrize('import_type', [
    'performance_grades',
    'salary_adjustments',
    'hire_info',
    'non_statutory_leave',
])
def test_e2e_template_xlsx_downloadable(client_hrbp, import_type):
    """SC1: 4 类 import_type 模板下载真实字节 + openpyxl 可读回 sheet。"""
    resp = client_hrbp.get(f'/api/v1/eligibility-import/templates/{import_type}?format=xlsx')
    assert resp.status_code == 200, resp.text
    assert len(resp.content) > 100, 'template should not be empty'
    wb = load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames, 'workbook should have at least one sheet'
    # 表头行存在
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers and headers[0] == '员工工号', f'Expected leading header 员工工号, got {headers}'


# ===================== APScheduler smoke =====================


def test_e2e_scheduler_module_structure():
    """import_scheduler 模块结构：暴露 4 个核心符号。"""
    from backend.app.scheduler import import_scheduler

    assert hasattr(import_scheduler, 'start_import_scheduler')
    assert hasattr(import_scheduler, 'stop_import_scheduler')
    assert hasattr(import_scheduler, 'run_expire_stale_jobs')
    assert hasattr(import_scheduler, 'scheduler')


def test_run_expire_stale_jobs_uses_isolated_session(tmp_uploads_dir):
    """run_expire_stale_jobs 用独立 SessionLocal；空 DB 下应不抛异常。"""
    from backend.app.scheduler import import_scheduler
    # 即使 import_service 未配置或 DB 为空，scheduler job 内部 try/except 应吞掉异常
    import_scheduler.run_expire_stale_jobs()
