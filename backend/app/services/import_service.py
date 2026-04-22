from __future__ import annotations

import csv
import hashlib
import io
import logging
import re
from collections.abc import Callable
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from fastapi import UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.core.config import get_settings
from backend.app.models.audit_log import AuditLog
from backend.app.models.certification import Certification
from backend.app.models.department import Department
from backend.app.models.employee import Employee
from backend.app.models.import_job import ImportJob
from backend.app.models.non_statutory_leave import NonStatutoryLeave
from backend.app.models.performance_record import PerformanceRecord
from backend.app.models.salary_adjustment_record import SalaryAdjustmentRecord
from backend.app.services.identity_binding_service import IdentityBindingService

logger = logging.getLogger(__name__)


class ImportService:
    # Phase 32 D-01: SUPPORTED_TYPES 扩展为 6 类（加 hire_info / non_statutory_leave）
    SUPPORTED_TYPES = {
        'employees', 'certifications', 'performance_grades', 'salary_adjustments',
        'hire_info', 'non_statutory_leave',
    }
    MAX_ROWS = 5000  # D-06: 单次导入最大行数

    # Phase 32-03 D-16 / D-17：ImportJob 状态常量
    # 持锁的状态（is_import_running 命中这些）
    _LOCKING_STATUSES = frozenset({'previewing', 'processing'})
    # 终态（不持锁，永远不会被 expire_stale 回收）
    _TERMINAL_STATUSES = frozenset({'completed', 'failed', 'partial', 'cancelled'})

    # Phase 32-03 D-09 / D-14: 业务键映射（4 类资格 import_type 使用，
    # 与 _import_* 内部上 upsert/查询逻辑保持一致；preview 阶段同文件冲突检测复用）
    # 与飞书同步 _sync_salary_adjustments_body line 947-952 业务键完全对齐
    _BUSINESS_KEYS = {
        'performance_grades': ['employee_no', 'year'],
        'salary_adjustments': ['employee_no', 'adjustment_date', 'adjustment_type'],
        'hire_info': ['employee_no'],
        'non_statutory_leave': ['employee_no', 'year'],
    }

    REQUIRED_COLUMNS = {
        'employees': ['employee_no', 'name', 'department', 'job_family', 'job_level'],
        'certifications': ['employee_no', 'certification_type', 'certification_stage', 'bonus_rate', 'issued_at'],
        'performance_grades': ['employee_no', 'year', 'grade'],
        'salary_adjustments': ['employee_no', 'adjustment_date', 'adjustment_type'],
        # Phase 32 D-02 / D-03
        'hire_info': ['employee_no', 'hire_date'],
        'non_statutory_leave': ['employee_no', 'year', 'total_days'],
    }
    COLUMN_ALIASES = {
        'employees': {
            '员工工号': 'employee_no',
            '员工姓名': 'name',
            '身份证号': 'id_card_no',
            '所属部门': 'department',
            '下属部门': 'sub_department',
            '所属公司': 'company',
            '岗位族': 'job_family',
            '岗位级别': 'job_level',
            '在职状态': 'status',
            '直属上级工号': 'manager_employee_no',
        },
        'certifications': {
            '员工工号': 'employee_no',
            '认证类型': 'certification_type',
            '认证阶段': 'certification_stage',
            '补贴比例': 'bonus_rate',
            '发证时间': 'issued_at',
            '到期时间': 'expires_at',
        },
        'performance_grades': {
            '员工工号': 'employee_no',
            '年度': 'year',
            '绩效等级': 'grade',
        },
        'salary_adjustments': {
            '员工工号': 'employee_no',
            '调薪日期': 'adjustment_date',
            '调薪类型': 'adjustment_type',
            '调薪金额': 'amount',
        },
        # Phase 32 D-02
        'hire_info': {
            '员工工号': 'employee_no',
            '入职日期': 'hire_date',
            '末次调薪日期': 'last_salary_adjustment_date',
            '历史调薪日期': 'last_salary_adjustment_date',  # 兼容飞书同步既有标签
        },
        # Phase 32 D-03
        'non_statutory_leave': {
            '员工工号': 'employee_no',
            '年度': 'year',
            '假期天数': 'total_days',
            '假期类型': 'leave_type',
        },
    }
    # D-09: 模板工号列预设文本格式的预填充行数（覆盖表头 + 示例 + 至少 100 行空行，
    # 避免 HR 新增数据时 Excel 重新识别列类型）。
    TEMPLATE_TEXT_PREFILL_ROWS = 105

    # D-09: 每种 import_type 下需要强制 cell.number_format = '@' 的列名（系统字段名）
    # D-10: employees 与 salary_adjustments 的 manager_employee_no 也属于工号类
    # D-09: certifications 的 certification_type / certification_stage 是短字符串，同样防误识别
    TEMPLATE_TEXT_COLUMNS = {
        'employees': ['employee_no', 'manager_employee_no'],
        'certifications': ['employee_no', 'certification_type', 'certification_stage'],
        'performance_grades': ['employee_no'],
        'salary_adjustments': ['employee_no'],
        # Phase 32 D-09
        'hire_info': ['employee_no'],
        'non_statutory_leave': ['employee_no'],
    }

    # D-07: 检测 Excel 工号列被错识别为数字后 pandas dtype=str 得到 '1234.0' 形态
    _LEADING_ZERO_LOST_PATTERN = re.compile(r'^\d+\.0$')

    # D-07: 需要跑 format sanity check 的「关键业务键」列（系统字段名）
    # 跨所有 import_type 适用；per-type 通过 REQUIRED_COLUMNS 进一步过滤
    _EMPLOYEE_NO_KEY_COLUMNS = frozenset({'employee_no', 'manager_employee_no'})

    COLUMN_LABELS = {
        'employee_no': '员工工号',
        'name': '员工姓名',
        'id_card_no': '身份证号',
        'department': '所属部门',
        'sub_department': '下属部门',
        'company': '所属公司',
        'job_family': '岗位族',
        'job_level': '岗位级别',
        'status': '在职状态',
        'manager_employee_no': '直属上级工号',
        'certification_type': '认证类型',
        'certification_stage': '认证阶段',
        'bonus_rate': '补贴比例',
        'issued_at': '发证时间',
        'expires_at': '到期时间',
        'year': '年度',
        'grade': '绩效等级',
        'adjustment_date': '调薪日期',
        'adjustment_type': '调薪类型',
        'amount': '调薪金额',
        # Phase 32
        'hire_date': '入职日期',
        'last_salary_adjustment_date': '末次调薪日期',
        'total_days': '假期天数',
        'leave_type': '假期类型',
    }

    def __init__(
        self,
        db: Session,
        *,
        operator_id: str | None = None,
        operator_role: str | None = None,
    ):
        self.db = db
        self._operator_id = operator_id
        self._operator_role = operator_role

    def _label_for_column(self, column_name: str) -> str:
        return self.COLUMN_LABELS.get(column_name, column_name)

    def _localize_error_message(self, message: str) -> str:
        mapping = {
            'Unsupported import type.': '暂不支持这个导入类型。',
            'Uploaded file is empty.': '上传文件为空，请重新选择文件后再导入。',
            'Import job not found.': '未找到这条导入任务记录。',
            'Unsupported file format. Please upload CSV for now.': '当前只支持导入 CSV 文件，请先把文件另存为 CSV 后再试。',
            'Excel import requires openpyxl in the current environment. Please upload CSV for now.': '当前环境暂不支持直接读取 Excel，请先另存为 CSV 后再导入。',
            'CSV 文件读取失败。': '文件读取失败，请检查文件内容后重试。',
            'Required employee fields cannot be empty.': '员工工号、员工姓名、所属部门、岗位族、岗位级别不能为空。',
            'Employee imported.': '导入成功。',
            'Certification imported.': '认证信息导入成功。',
            'Invalid certification date or bonus rate.': '认证信息格式不正确，请检查发证时间、到期时间或补贴比例。',
            'This ID card number is already used by another employee profile.': '该身份证号已被其他员工档案占用。',
            'This ID card number is already used by another account.': '该身份证号已被其他平台账号占用。',
            'This employee profile is already bound to another account.': '该员工档案已绑定其他平台账号。',
        }
        if message in mapping:
            return mapping[message]
        if message.startswith('Department ') and message.endswith(' was not found.'):
            department_name = message[len('Department '):-len(' was not found.')]
            return f'部门"{department_name}"未创建，请先到部门管理中新增。'
        if message.startswith('Department ') and message.endswith(' is inactive.'):
            department_name = message[len('Department '):-len(' is inactive.')]
            return f'部门"{department_name}"已停用，请启用后再导入。'
        return message

    def _resolve_department_name(self, department_name: str) -> str:
        normalized_name = department_name.strip()
        department = self.db.scalar(select(Department).where(Department.name == normalized_name))
        if department is None:
            raise ValueError(f'Department {normalized_name} was not found.')
        if department.status != 'active':
            raise ValueError(f'Department {normalized_name} is inactive.')
        return department.name

    def list_jobs(self) -> list[ImportJob]:
        query = select(ImportJob).order_by(ImportJob.created_at.desc())
        return list(self.db.scalars(query))

    def get_job(self, job_id: str) -> ImportJob | None:
        return self.db.get(ImportJob, job_id)

    def delete_job(self, job_id: str) -> str:
        job = self.get_job(job_id)
        if job is None:
            raise ValueError('Import job not found.')
        self.db.delete(job)
        self.db.commit()
        return job_id

    def bulk_delete_jobs(self, job_ids: list[str]) -> list[str]:
        deleted_ids: list[str] = []
        for job_id in job_ids:
            job = self.get_job(job_id)
            if job is None:
                continue
            self.db.delete(job)
            deleted_ids.append(job_id)
        self.db.commit()
        return deleted_ids

    def run_import(
        self,
        *,
        import_type: str,
        upload: UploadFile,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
        progress_callback: Callable[[int, int, int], None] | None = None,
    ) -> ImportJob:
        normalized_type = import_type.strip().lower()
        if normalized_type not in self.SUPPORTED_TYPES:
            raise ValueError(self._localize_error_message('Unsupported import type.'))
        # Phase 32 D-12: overwrite_mode 仅接受 merge / replace
        normalized_mode = (overwrite_mode or 'merge').strip().lower()
        if normalized_mode not in ('merge', 'replace'):
            raise ValueError(f'overwrite_mode 必须是 merge 或 replace，当前值：{overwrite_mode!r}')
        file_name = upload.filename or f'{normalized_type}.csv'
        raw_bytes = upload.file.read()
        if not raw_bytes:
            raise ValueError(self._localize_error_message('Uploaded file is empty.'))

        job = ImportJob(
            file_name=file_name,
            import_type=normalized_type,
            status='processing',
            total_rows=0,
            success_rows=0,
            failed_rows=0,
            result_summary={'rows': []},
            overwrite_mode=normalized_mode,
            actor_id=self._operator_id,
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)

        try:
            dataframe = self._load_table(file_name, raw_bytes)
            # D-06: 单次导入不能超过 MAX_ROWS 行
            if len(dataframe) > self.MAX_ROWS:
                raise ValueError(f'单次导入不能超过 {self.MAX_ROWS} 行，请分批导入。')
            if progress_callback:
                progress_callback(0, len(dataframe), 0)
            row_results = self._dispatch_import(normalized_type, dataframe, overwrite_mode=normalized_mode)
            job.total_rows = len(row_results)
            job.success_rows = sum(1 for item in row_results if item['status'] == 'success')
            job.failed_rows = sum(1 for item in row_results if item['status'] == 'failed')
            if progress_callback:
                progress_callback(job.total_rows, job.total_rows, job.failed_rows)
            job.result_summary = {
                'rows': row_results,
                'supported_types': sorted(self.SUPPORTED_TYPES),
            }
            if progress_callback:
                progress_callback(job.total_rows, job.total_rows, job.failed_rows)
            # Status: 'partial' when mixed, 'failed' when all fail, 'completed' when all succeed
            if job.failed_rows == 0:
                job.status = 'completed'
            elif job.success_rows == 0:
                job.status = 'failed'
            else:
                job.status = 'partial'
        except Exception as exc:
            job.status = 'failed'
            job.result_summary = {
                'rows': [],
                'error': self._localize_error_message(str(exc)),
                'supported_types': sorted(self.SUPPORTED_TYPES),
            }
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def build_template(self, import_type: str) -> tuple[str, bytes, str]:
        normalized_type = import_type.strip().lower()
        if normalized_type not in self.SUPPORTED_TYPES:
            raise ValueError(self._localize_error_message('Unsupported import type.'))
        output = io.StringIO()
        writer = csv.writer(output)
        if normalized_type == 'employees':
            writer.writerow(['员工工号', '员工姓名', '身份证号', '所属部门', '下属部门', '所属公司', '岗位族', '岗位级别', '在职状态', '直属上级工号'])
            writer.writerow(['EMP-1001', '张小明', '310101199001010123', '产品技术中心', '后端平台组', '示例科技', '平台研发', 'P5', 'active', ''])
        elif normalized_type == 'certifications':
            writer.writerow(['员工工号', '认证类型', '认证阶段', '补贴比例', '发证时间', '到期时间'])
            writer.writerow(['EMP-1001', 'ai_skill', 'advanced', '0.02', '2026-01-15T00:00:00+00:00', ''])
        elif normalized_type == 'performance_grades':
            writer.writerow(['员工工号', '年度', '绩效等级'])
            writer.writerow(['EMP-1001', '2025', 'B'])
        elif normalized_type == 'salary_adjustments':
            writer.writerow(['员工工号', '调薪日期', '调薪类型', '调薪金额'])
            writer.writerow(['EMP-1001', '2025-06-01', '年度调薪', '2000'])
        elif normalized_type == 'hire_info':
            writer.writerow(['员工工号', '入职日期', '末次调薪日期'])
            writer.writerow(['EMP-1001', '2026-01-15', '2025-06-01'])
        elif normalized_type == 'non_statutory_leave':
            writer.writerow(['员工工号', '年度', '假期天数', '假期类型'])
            writer.writerow(['EMP-1001', '2026', '10.5', '事假'])
        # UTF-8 BOM helps Excel on Windows open Chinese CSVs correctly.
        content = output.getvalue().encode('utf-8-sig')
        return f'{normalized_type}_template.csv', content, 'text/csv; charset=utf-8'

    def build_template_xlsx(self, import_type: str) -> tuple[str, bytes, str]:
        """Build an xlsx template file for the given import type (IMP-06, D-04)."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        normalized_type = import_type.strip().lower()
        if normalized_type not in self.SUPPORTED_TYPES:
            raise ValueError(self._localize_error_message('Unsupported import type.'))

        wb = Workbook()
        ws = wb.active
        ws.title = '导入模板'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        if normalized_type == 'employees':
            headers = ['员工工号', '员工姓名', '身份证号', '所属部门', '下属部门', '所属公司', '岗位族', '岗位级别', '在职状态', '直属上级工号']
            example = ['02651', '张小明', '310101199001010123', '产品技术中心', '后端平台组', '示例科技', '平台研发', 'P5', 'active', '02650']
        elif normalized_type == 'certifications':
            headers = ['员工工号', '认证类型', '认证阶段', '补贴比例', '发证时间', '到期时间']
            example = ['02651', 'ai_skill', 'advanced', '0.02', '2026-01-15T00:00:00+00:00', '']
        elif normalized_type == 'performance_grades':
            headers = ['员工工号', '年度', '绩效等级']
            example = ['02651', '2025', 'B']
        elif normalized_type == 'salary_adjustments':
            headers = ['员工工号', '调薪日期', '调薪类型', '调薪金额']
            example = ['02651', '2025-06-01', '年度调薪', '2000']
        elif normalized_type == 'hire_info':
            # Phase 32 D-02
            headers = ['员工工号', '入职日期', '末次调薪日期']
            example = ['02651', '2026-01-15', '2025-06-01']
        elif normalized_type == 'non_statutory_leave':
            # Phase 32 D-03
            headers = ['员工工号', '年度', '假期天数', '假期类型']
            example = ['02651', 2026, 10.5, '事假']
        else:
            headers = []
            example = []

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for col_idx, value in enumerate(example, start=1):
            ws.cell(row=2, column=col_idx, value=value)

        # Set column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 20

        # Properly set column widths for all columns
        from openpyxl.utils import get_column_letter  # noqa: F401 (供需要时使用)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        # D-09: 对文本类列（工号、认证类型/阶段等）预填充 number_format='@'，
        # 覆盖表头 + 示例 + 103 行空行 = 105 行，确保 HR 录入时 Excel 不会重新识别列类型
        text_column_system_names = self.TEMPLATE_TEXT_COLUMNS.get(normalized_type, [])
        alias_map = self.COLUMN_ALIASES.get(normalized_type, {})
        # system_field -> chinese_header 反向映射
        reverse_alias = {v: k for k, v in alias_map.items()}
        for system_name in text_column_system_names:
            header_label = reverse_alias.get(system_name)
            if header_label is None:
                continue
            try:
                col_idx = headers.index(header_label) + 1
            except ValueError:
                continue
            for row_idx in range(1, self.TEMPLATE_TEXT_PREFILL_ROWS + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '@'

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return f'{normalized_type}_template.xlsx', content, media_type

    def build_export_report(self, job: ImportJob) -> tuple[str, bytes, str]:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['行号', '结果', '提示'])
        rows = job.result_summary.get('rows', []) if isinstance(job.result_summary, dict) else []
        for item in rows:
            writer.writerow([item.get('row_index', ''), item.get('status', ''), item.get('message', '')])
        if not rows:
            writer.writerow([
                '',
                'failed',
                job.result_summary.get('error', '当前没有可导出的行级结果。') if isinstance(job.result_summary, dict) else '当前没有可导出的行级结果。',
            ])
        content = output.getvalue().encode('utf-8-sig')
        return f'{job.import_type}_{job.id}_report.csv', content, 'text/csv; charset=utf-8'

    def build_export_report_xlsx(self, job: ImportJob) -> tuple[str, bytes, str]:
        """Build an xlsx error report for a completed import job (D-02)."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = '导入结果报告'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center')

        headers = ['行号', '结果', '错误字段', '错误原因']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        rows = job.result_summary.get('rows', []) if isinstance(job.result_summary, dict) else []
        # Only include failed rows in the report
        row_num = 2
        for item in rows:
            if item.get('status') != 'failed':
                continue
            ws.cell(row=row_num, column=1, value=item.get('row_index', ''))
            ws.cell(row=row_num, column=2, value=item.get('status', ''))
            ws.cell(row=row_num, column=3, value=item.get('error_column', ''))
            ws.cell(row=row_num, column=4, value=item.get('message', ''))
            row_num += 1

        if row_num == 2:
            # No failed rows — write a summary row
            error_msg = ''
            if isinstance(job.result_summary, dict):
                error_msg = job.result_summary.get('error', '当前没有失败行。')
            else:
                error_msg = '当前没有失败行。'
            ws.cell(row=2, column=1, value='')
            ws.cell(row=2, column=2, value='info')
            ws.cell(row=2, column=3, value='')
            ws.cell(row=2, column=4, value=error_msg)

        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return f'{job.import_type}_{job.id}_report.xlsx', content, media_type

    @staticmethod
    def _parse_excel_date(value):
        """Phase 32 Pitfall 1：解析 Excel 单元格日期值。

        ImportService 用 ``pd.read_excel(..., dtype=str)`` 读入数据，因此 Excel 序列号日期
        会被转成字符串（如 '45292'）。本 helper 处理三类输入：

        1. ``None`` / 空字符串 → 返回 ``None``
        2. 已是 ``date`` / ``datetime`` 实例 → 直接返回 date
        3. 数值 / 5-6 位纯数字字符串 → 视为 Excel 序列号（origin='1899-12-30'）
        4. 其余字符串 → 走 ``pd.to_datetime`` 通用解析（支持 ISO 8601 等）

        无法解析时抛 ``ValueError``，由调用方捕获并写入 row failed 结果。
        """
        from datetime import date as _date, datetime as _datetime
        if value is None:
            return None
        if isinstance(value, _datetime):
            return value.date()
        if isinstance(value, _date):
            return value
        if isinstance(value, (int, float)):
            return pd.to_datetime(value, unit='D', origin='1899-12-30').date()
        s = str(value).strip()
        if not s:
            return None
        if s.isdigit() and 5 <= len(s) <= 6:
            return pd.to_datetime(int(s), unit='D', origin='1899-12-30').date()
        try:
            return pd.to_datetime(s).date()
        except Exception as exc:
            raise ValueError(f'日期格式无效: {value!r} ({exc})') from exc

    def _load_table(self, file_name: str, raw_bytes: bytes) -> pd.DataFrame:
        suffix = file_name.lower().rsplit('.', 1)[-1] if '.' in file_name else ''
        if suffix == 'csv':
            last_error: Exception | None = None
            for encoding in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk'):
                try:
                    return pd.read_csv(io.BytesIO(raw_bytes), encoding=encoding, dtype=str).fillna('')
                except UnicodeDecodeError as exc:
                    last_error = exc
            if last_error is not None:
                raise ValueError('CSV 文件编码格式不正确，请使用"CSV UTF-8"或 GBK/GB18030 编码重新保存后再导入。') from last_error
            raise ValueError(self._localize_error_message('CSV 文件读取失败。'))
        if suffix in {'xlsx', 'xls'}:
            # IMP-04: xlsx/xls support via openpyxl
            return pd.read_excel(io.BytesIO(raw_bytes), engine='openpyxl', dtype=str).fillna('')
        raise ValueError(self._localize_error_message('Unsupported file format. Please upload CSV for now.'))

    def _dispatch_import(
        self,
        import_type: str,
        dataframe: pd.DataFrame,
        *,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
    ) -> list[dict[str, object]]:
        # row_index 语义（全局约定）：pandas DataFrame 0-based index 加 1 = 数据行号（不含 Excel 表头）。
        # HR 对照 Excel 时需加 1（row_index=2 对应 Excel 第 3 行）。Phase 30 内不扩大改动。
        dataframe = self._normalize_columns(import_type, dataframe)
        required = self.REQUIRED_COLUMNS[import_type]
        missing = [column for column in required if column not in dataframe.columns]
        if missing:
            missing_labels = '、'.join(self._label_for_column(column) for column in missing)
            raise ValueError(f'缺少必填列：{missing_labels}。请重新下载最新模板后填写。')

        # D-07: format sanity check —— 在派发到具体 _import_* 前检测前导零丢失
        bad_rows = self._detect_leading_zero_loss_rows(dataframe)
        results: list[dict[str, object]] = []
        if bad_rows:
            for pandas_idx in sorted(bad_rows.keys()):
                row_index, col_name = bad_rows[pandas_idx]
                col_label = self._label_for_column(col_name)
                results.append({
                    'row_index': row_index,
                    'status': 'failed',
                    'error_column': col_label,
                    'message': (
                        f'第 {row_index} 行{col_label}列格式异常（疑似丢失前导零）。'
                        f'请在 Excel 中将该列改为「文本」格式后重新上传，或从系统重新下载最新模板。'
                    ),
                })
            # 用 bool mask 丢弃坏行，**保留原 pandas index**（禁止 reset_index）
            bad_mask = dataframe.index.isin(bad_rows.keys())
            dataframe = dataframe.loc[~bad_mask].copy()

        if import_type == 'employees':
            results.extend(self._import_employees(dataframe))
        elif import_type == 'performance_grades':
            results.extend(self._import_performance_grades(dataframe, overwrite_mode=overwrite_mode))
        elif import_type == 'salary_adjustments':
            results.extend(self._import_salary_adjustments(dataframe, overwrite_mode=overwrite_mode))
        elif import_type == 'hire_info':
            # Phase 32 D-02
            results.extend(self._import_hire_info(dataframe, overwrite_mode=overwrite_mode))
        elif import_type == 'non_statutory_leave':
            # Phase 32 D-03
            results.extend(self._import_non_statutory_leave(dataframe, overwrite_mode=overwrite_mode))
        else:
            results.extend(self._import_certifications(dataframe))
        return results

    def _normalize_columns(self, import_type: str, dataframe: pd.DataFrame) -> pd.DataFrame:
        alias_map = self.COLUMN_ALIASES.get(import_type, {})
        normalized_columns: dict[str, str] = {}
        for column in dataframe.columns:
            stripped = str(column).strip()
            normalized_columns[column] = alias_map.get(stripped, stripped)
        return dataframe.rename(columns=normalized_columns)

    def _detect_error_column(self, exc: Exception) -> str:
        """Detect which column caused the employee import error based on exception message."""
        msg = str(exc)
        if '不能为空' in msg or 'cannot be empty' in msg.lower():
            return '必填字段'
        if '部门' in msg and ('未创建' in msg or 'not found' in msg.lower() or '停用' in msg):
            return '所属部门'
        if '身份证号' in msg or 'id_card' in msg.lower() or 'ID card' in msg:
            return '身份证号'
        if '员工工号' in msg or 'employee_no' in msg.lower():
            return '员工工号'
        return ''

    def _detect_cert_error_column(self, exc: Exception) -> str:
        """Detect which column caused the certification import error based on exception message."""
        msg = str(exc)
        if '员工工号' in msg or '未找到' in msg:
            return '员工工号'
        if '日期' in msg or '时间' in msg or 'datetime' in msg.lower() or 'date' in msg.lower():
            return '发证时间/到期时间'
        if '补贴' in msg or 'bonus_rate' in msg.lower() or 'float' in msg.lower():
            return '补贴比例'
        return ''

    @classmethod
    def _detect_leading_zero_loss_rows(cls, dataframe: pd.DataFrame) -> dict[int, tuple[int, str]]:
        """D-07: 扫描 dataframe 中的关键业务键列，返回 {pandas_idx: (row_index, column_name)} 映射。

        row_index 语义：pandas 0-based index 加 1 = 数据行号（不含 Excel 表头）。
        **与既有 _import_* 方法的 row_index 口径完全一致**（HR 对照 Excel 时需加 1 跳过表头行）。
        值匹配 '^\\d+\\.0$' 视为「工号列格式异常（疑似丢失前导零）」。
        """
        bad_rows: dict[int, tuple[int, str]] = {}
        for col_name in cls._EMPLOYEE_NO_KEY_COLUMNS:
            if col_name not in dataframe.columns:
                continue
            for pandas_idx, value in dataframe[col_name].items():
                if isinstance(value, str) and cls._LEADING_ZERO_LOST_PATTERN.match(value.strip()):
                    row_index = int(pandas_idx) + 1  # 与 _import_* 口径一致
                    if pandas_idx not in bad_rows:
                        bad_rows[pandas_idx] = (row_index, col_name)
        return bad_rows

    def _import_employees(self, dataframe: pd.DataFrame) -> list[dict[str, object]]:
        results: list[dict[str, object]] = []
        staged_rows: list[tuple[Employee, str | None]] = []
        identity_service = IdentityBindingService(self.db)
        has_company_column = 'company' in dataframe.columns
        for index, row in dataframe.iterrows():
            try:
                with self.db.begin_nested():  # SAVEPOINT
                    employee_no = str(row['employee_no']).strip()
                    name = str(row['name']).strip()
                    id_card_no = str(row['id_card_no']).strip() if 'id_card_no' in dataframe.columns else ''
                    department = str(row['department']).strip()
                    sub_department = str(row['sub_department']).strip() if 'sub_department' in dataframe.columns else ''
                    company = str(row['company']).strip() if has_company_column else ''
                    job_family = str(row['job_family']).strip()
                    job_level = str(row['job_level']).strip()
                    status_val = str(row['status']).strip() or 'active'
                    manager_no = str(row['manager_employee_no']).strip() if 'manager_employee_no' in dataframe.columns else ''
                    if not all([employee_no, name, department, job_family, job_level]):
                        raise ValueError(self._localize_error_message('Required employee fields cannot be empty.'))

                    employee = self.db.scalar(select(Employee).where(Employee.employee_no == employee_no))
                    # employee_no has DB-level unique constraint + index, upsert is reliable
                    department = self._resolve_department_name(department)
                    normalized_id_card_no = identity_service.ensure_employee_id_card_available(
                        id_card_no or None,
                        employee_id=employee.id if employee is not None else None,
                    )

                    if employee is None:
                        employee = Employee(
                            employee_no=employee_no,
                            name=name,
                            id_card_no=normalized_id_card_no,
                            department=department,
                            sub_department=sub_department or None,
                            company=company or None,
                            job_family=job_family,
                            job_level=job_level,
                            status=status_val,
                        )
                    else:
                        # D-03/IMP-05: Record old values for audit log before updating
                        old_values = {
                            'name': employee.name,
                            'department': employee.department,
                            'job_family': employee.job_family,
                            'job_level': employee.job_level,
                            'status': employee.status,
                        }
                        if has_company_column:
                            old_values['company'] = employee.company
                        employee.name = name
                        employee.id_card_no = normalized_id_card_no
                        employee.department = department
                        employee.sub_department = sub_department or None
                        if has_company_column:
                            employee.company = company or None
                        employee.job_family = job_family
                        employee.job_level = job_level
                        employee.status = status_val
                        new_value = {
                            'name': name,
                            'department': department,
                            'job_family': job_family,
                            'job_level': job_level,
                            'status': status_val,
                        }
                        if has_company_column:
                            new_value['company'] = employee.company
                        # Write audit log for employee update
                        audit_entry = AuditLog(
                            operator_id=self._operator_id,
                            action='employee_import_update',
                            target_type='employee',
                            target_id=employee.id,
                            detail={
                                'old_value': old_values,
                                'new_value': new_value,
                                'operator_role': self._operator_role,
                            },
                        )
                        self.db.add(audit_entry)

                    self.db.add(employee)
                    self.db.flush()

                    bind_warning: str | None = None
                    try:
                        identity_service.auto_bind_user_and_employee(employee=employee)
                    except ValueError as exc:
                        bind_warning = self._localize_error_message(str(exc))

                staged_rows.append((employee, manager_no or None))
                success_message = self._localize_error_message('Employee imported.')
                if bind_warning:
                    success_message = f'{success_message} 但自动绑定账号失败：{bind_warning}'
                results.append({'row_index': int(index) + 1, 'status': 'success', 'message': success_message})
            except Exception as exc:
                self.db.expire_all()  # Clean stale objects after SAVEPOINT rollback
                results.append({
                    'row_index': int(index) + 1,
                    'status': 'failed',
                    'message': self._localize_error_message(str(exc)),
                    'error_column': self._detect_error_column(exc),
                })

        # Second pass: manager binding with SAVEPOINT per row
        for employee, manager_no in staged_rows:
            try:
                with self.db.begin_nested():  # SAVEPOINT for manager binding
                    if not manager_no:
                        employee.manager_id = None
                        self.db.add(employee)
                    else:
                        manager = self.db.scalar(select(Employee).where(Employee.employee_no == manager_no))
                        if manager is None:
                            employee.manager_id = None
                            results.append({
                                'row_index': None,
                                'status': 'failed',
                                'message': f'未找到直属上级工号"{manager_no}"，员工 {employee.employee_no} 已跳过上级绑定。',
                                'error_column': '直属上级工号',
                            })
                        else:
                            employee.manager_id = manager.id
                            self.db.add(employee)
            except Exception as exc:
                self.db.expire_all()  # Clean stale objects after SAVEPOINT rollback
                results.append({
                    'row_index': None,
                    'status': 'failed',
                    'message': f'上级绑定失败：{exc!s}',
                    'error_column': '直属上级工号',
                })

        self.db.commit()
        return results

    def _import_certifications(self, dataframe: pd.DataFrame) -> list[dict[str, object]]:
        results: list[dict[str, object]] = []
        for index, row in dataframe.iterrows():
            try:
                with self.db.begin_nested():  # SAVEPOINT
                    employee_no = str(row['employee_no']).strip()
                    employee = self.db.scalar(select(Employee).where(Employee.employee_no == employee_no))
                    if employee is None:
                        raise ValueError(f'未找到员工工号"{employee_no}"，请先导入员工档案。')

                    try:
                        issued_at = pd.to_datetime(row['issued_at'], utc=True).to_pydatetime()
                        expires_value = str(row['expires_at']).strip() if 'expires_at' in dataframe.columns else ''
                        expires_at = pd.to_datetime(expires_value, utc=True).to_pydatetime() if expires_value else None
                        bonus_rate = float(row['bonus_rate'])
                    except Exception:
                        raise ValueError(self._localize_error_message('Invalid certification date or bonus rate.'))

                    certification_type = str(row['certification_type']).strip()
                    # Upsert on (employee_id, certification_type)
                    existing = self.db.scalar(
                        select(Certification).where(
                            Certification.employee_id == employee.id,
                            Certification.certification_type == certification_type,
                        )
                    )
                    if existing is not None:
                        existing.certification_stage = str(row['certification_stage']).strip()
                        existing.bonus_rate = bonus_rate
                        existing.issued_at = issued_at
                        existing.expires_at = expires_at
                        self.db.add(existing)
                    else:
                        certification = Certification(
                            employee_id=employee.id,
                            certification_type=certification_type,
                            certification_stage=str(row['certification_stage']).strip(),
                            bonus_rate=bonus_rate,
                            issued_at=issued_at,
                            expires_at=expires_at,
                        )
                        self.db.add(certification)
                    self.db.flush()
                results.append({'row_index': int(index) + 1, 'status': 'success', 'message': '认证导入成功。'})
            except Exception as exc:
                self.db.expire_all()  # Clean stale objects after SAVEPOINT rollback
                results.append({
                    'row_index': int(index) + 1,
                    'status': 'failed',
                    'message': self._localize_error_message(str(exc)),
                    'error_column': self._detect_cert_error_column(exc),
                })
        self.db.commit()
        return results

    # ------------------------------------------------------------------
    # Performance grades import (D-07/ELIG-09)
    # ------------------------------------------------------------------

    def _import_performance_grades(
        self,
        dataframe: pd.DataFrame,
        *,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
    ) -> list[dict[str, object]]:
        # 业务键定义在类常量 _BUSINESS_KEYS（Phase 32-03 已抽取）
        # preview 阶段同文件冲突检测复用同一份 key 列表
        results: list[dict[str, object]] = []
        for index, row in dataframe.iterrows():
            row_no = int(index) + 1
            try:
                with self.db.begin_nested():  # SAVEPOINT
                    employee_no = str(row['employee_no']).strip()
                    employee = self.db.scalar(select(Employee).where(Employee.employee_no == employee_no))
                    if employee is None:
                        raise ValueError(f'未找到员工工号 "{employee_no}"')

                    year = int(row['year'])

                    # Phase 32 D-12: grade 是必填字段
                    raw_grade = row.get('grade')
                    grade_str = str(raw_grade).strip() if raw_grade is not None else ''
                    if not grade_str:
                        if overwrite_mode == 'replace':
                            raise ValueError('替换模式下绩效等级为必填字段，不允许清空')
                        # merge：跳过不更新，作为 no_change 处理
                        results.append({
                            'row_index': row_no,
                            'row': row_no,
                            'status': 'success',
                            'employee_no': employee_no,
                            'action': 'no_change',
                            'message': '绩效等级为空，merge 模式下保留旧值。',
                        })
                        continue

                    grade = grade_str.upper()
                    if grade not in ('A', 'B', 'C', 'D', 'E'):
                        raise ValueError(f'绩效等级 "{grade}" 不合法，请填写 A/B/C/D/E')

                    # Idempotent upsert on (employee_id, year)
                    existing = self.db.scalar(
                        select(PerformanceRecord).where(
                            PerformanceRecord.employee_id == employee.id,
                            PerformanceRecord.year == year,
                        )
                    )
                    if existing is not None:
                        old_grade = existing.grade
                        existing.grade = grade
                        existing.source = 'excel'
                        # Phase 34 B-1 + D-08：刷新部门快照（None 时也写 None）
                        existing.department_snapshot = employee.department
                        self.db.add(existing)
                        action = 'update' if old_grade != grade else 'no_change'
                    else:
                        record = PerformanceRecord(
                            employee_id=employee.id,
                            employee_no=employee_no,
                            year=year,
                            grade=grade,
                            source='excel',
                            # Phase 34 B-1 + D-08：录入时部门快照（None 时也写 None）
                            department_snapshot=employee.department,
                        )
                        self.db.add(record)
                        action = 'insert'
                    self.db.flush()
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'success',
                    'employee_no': employee_no,
                    'action': action,
                    'message': '绩效记录导入成功。' if action != 'no_change' else '无字段变化，跳过更新。',
                })
            except Exception as exc:
                self.db.expire_all()
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'message': str(exc),
                    'error': str(exc),
                    'error_column': self._detect_perf_error_column(exc),
                })
        self.db.commit()
        return results

    def _detect_perf_error_column(self, exc: Exception) -> str:
        msg = str(exc)
        if '员工工号' in msg or '未找到' in msg:
            return '员工工号'
        if '绩效等级' in msg:
            return '绩效等级'
        if '年度' in msg or 'year' in msg.lower():
            return '年度'
        return ''

    # ------------------------------------------------------------------
    # Salary adjustments import (D-07/ELIG-09)
    # ------------------------------------------------------------------

    _ADJ_TYPE_MAP: dict[str, str] = {
        '转正调薪': 'probation',
        '年度调薪': 'annual',
        '专项调薪': 'special',
    }

    def _import_salary_adjustments(
        self,
        dataframe: pd.DataFrame,
        *,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
    ) -> list[dict[str, object]]:
        """Phase 32 D-14 / Pitfall 4：改 upsert by (employee_id, adjustment_date, adjustment_type)，
        与飞书同步 _sync_salary_adjustments_body 业务键对齐。

        amount 是可选字段：
          - merge：空保留旧值
          - replace：空清空旧值
        """
        # 业务键定义在类常量 _BUSINESS_KEYS（Phase 32-03 已抽取）
        # preview 阶段同文件冲突检测复用同一份 key 列表
        results: list[dict[str, object]] = []

        for index, row in dataframe.iterrows():
            row_no = int(index) + 1
            try:
                with self.db.begin_nested():  # SAVEPOINT
                    employee_no = str(row['employee_no']).strip()
                    employee = self.db.scalar(
                        select(Employee).where(Employee.employee_no == employee_no)
                    )
                    if employee is None:
                        raise ValueError(f'未找到员工工号 "{employee_no}"')

                    # Parse adjustment_date with Phase 32 _parse_excel_date helper
                    raw_date = row['adjustment_date']
                    try:
                        adjustment_date = self._parse_excel_date(raw_date)
                    except ValueError:
                        raise ValueError(f'调薪日期格式无效: "{raw_date}"')
                    if adjustment_date is None:
                        raise ValueError('调薪日期为必填字段')

                    # Parse adjustment_type with Chinese-to-code mapping
                    raw_type = str(row['adjustment_type']).strip()
                    if not raw_type:
                        raise ValueError('调薪类型为必填字段')
                    adj_type = self._ADJ_TYPE_MAP.get(raw_type, raw_type.lower())
                    if adj_type not in ('probation', 'annual', 'special'):
                        raise ValueError(
                            f'调薪类型 "{raw_type}" 不合法，请填写 转正调薪/年度调薪/专项调薪 或 probation/annual/special'
                        )

                    # Parse amount (optional)
                    amount: Decimal | None = None
                    if 'amount' in dataframe.columns:
                        raw_amount = row['amount']
                        if pd.notna(raw_amount) and str(raw_amount).strip():
                            try:
                                amount = Decimal(str(raw_amount).strip())
                            except InvalidOperation:
                                raise ValueError(f'调薪金额格式无效: "{raw_amount}"')

                    # D-14: upsert by (employee_id, adjustment_date, adjustment_type)
                    existing = self.db.scalar(
                        select(SalaryAdjustmentRecord).where(
                            SalaryAdjustmentRecord.employee_id == employee.id,
                            SalaryAdjustmentRecord.adjustment_date == adjustment_date,
                            SalaryAdjustmentRecord.adjustment_type == adj_type,
                        )
                    )

                    if existing is None:
                        record = SalaryAdjustmentRecord(
                            employee_id=employee.id,
                            employee_no=employee_no,
                            adjustment_date=adjustment_date,
                            adjustment_type=adj_type,
                            amount=amount,
                            source='excel',
                        )
                        self.db.add(record)
                        self.db.flush()
                        results.append({
                            'row_index': row_no,
                            'row': row_no,
                            'status': 'success',
                            'employee_no': employee_no,
                            'action': 'insert',
                            'message': '调薪记录新增成功。',
                            'fields': {
                                'amount': {'old': None, 'new': str(amount) if amount is not None else None},
                            },
                        })
                    else:
                        old_amount = existing.amount
                        did_change = False
                        # amount 是可选字段
                        #   merge：空保留 / 非空覆盖
                        #   replace：空清空 / 非空覆盖
                        if amount is None:
                            if overwrite_mode == 'replace' and old_amount is not None:
                                existing.amount = None
                                did_change = True
                        else:
                            if amount != old_amount:
                                existing.amount = amount
                                did_change = True
                        existing.source = 'excel'
                        self.db.flush()
                        results.append({
                            'row_index': row_no,
                            'row': row_no,
                            'status': 'success',
                            'employee_no': employee_no,
                            'action': 'update' if did_change else 'no_change',
                            'message': '调薪记录更新成功。' if did_change else '无字段变化，跳过更新。',
                            'fields': {
                                'amount': {
                                    'old': str(old_amount) if old_amount is not None else None,
                                    'new': str(amount) if amount is not None else (
                                        None if overwrite_mode == 'replace' else (
                                            str(old_amount) if old_amount is not None else None
                                        )
                                    ),
                                },
                            },
                        })
            except Exception as exc:
                self.db.expire_all()
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'message': str(exc),
                    'error': str(exc),
                    'error_column': self._detect_adj_error_column(exc),
                })
        self.db.commit()
        return results

    def _detect_adj_error_column(self, exc: Exception) -> str:
        msg = str(exc)
        if '员工工号' in msg or '未找到' in msg:
            return '员工工号'
        if '调薪日期' in msg:
            return '调薪日期'
        if '调薪类型' in msg:
            return '调薪类型'
        if '调薪金额' in msg:
            return '调薪金额'
        return ''

    # ------------------------------------------------------------------
    # Phase 32 D-02: Hire info import (IMPORT-01)
    # ------------------------------------------------------------------

    def _import_hire_info(
        self,
        dataframe: pd.DataFrame,
        *,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
    ) -> list[dict[str, object]]:
        """D-02 + IMPORT-05：导入入职信息到 Employee 表，业务键 = employee_id。

        merge 模式：空字段保留旧值。
        replace 模式：必填字段（hire_date）为空 → row failed；时间戳类的可选字段
        last_salary_adjustment_date 也保留旧值（CONTEXT D-子提示：时间戳无 NULL 语义）。
        """
        # 业务键定义在类常量 _BUSINESS_KEYS（Phase 32-03 已抽取）
        # preview 阶段同文件冲突检测复用同一份 key 列表
        results: list[dict[str, object]] = []

        # 内联自查 employee_no → Employee（决议 Warning 2：不引入 IdentityBindingService 跨服务方法）
        employee_nos = [
            str(v).strip()
            for v in dataframe.get('employee_no', [])
            if v is not None and str(v).strip()
        ]
        emp_rows = (
            self.db.execute(
                select(Employee).where(Employee.employee_no.in_(employee_nos))
            ).scalars().all()
            if employee_nos else []
        )
        emp_map: dict[str, Employee] = {e.employee_no: e for e in emp_rows}

        for index, row in dataframe.iterrows():
            row_no = int(index) + 1
            emp_no = str(row.get('employee_no', '')).strip()
            if not emp_no:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': '',
                    'error': '员工工号为必填字段',
                    'message': '员工工号为必填字段',
                    'error_column': '员工工号',
                })
                continue

            employee = emp_map.get(emp_no)
            if employee is None:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': f'未找到员工工号 "{emp_no}"',
                    'message': f'未找到员工工号 "{emp_no}"',
                    'error_column': '员工工号',
                })
                continue

            raw_hire = row.get('hire_date')
            raw_last_adj = row.get('last_salary_adjustment_date') if 'last_salary_adjustment_date' in dataframe.columns else None

            try:
                new_hire = self._parse_excel_date(raw_hire)
            except ValueError as exc:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': f'入职日期格式无效: {raw_hire!r}',
                    'message': f'入职日期格式无效: {raw_hire!r}',
                    'error_column': '入职日期',
                })
                continue

            try:
                new_last_adj = self._parse_excel_date(raw_last_adj)
            except ValueError as exc:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': f'末次调薪日期格式无效: {raw_last_adj!r}',
                    'message': f'末次调薪日期格式无效: {raw_last_adj!r}',
                    'error_column': '末次调薪日期',
                })
                continue

            # D-12：replace 模式必填字段 hire_date 空 → failed（不允许清空）
            if new_hire is None and overwrite_mode == 'replace':
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': '替换模式下入职日期为必填字段，不允许清空',
                    'message': '替换模式下入职日期为必填字段，不允许清空',
                    'error_column': '入职日期',
                })
                continue

            old_hire = employee.hire_date
            old_last_adj = employee.last_salary_adjustment_date
            did_change = False

            if new_hire is not None and new_hire != old_hire:
                employee.hire_date = new_hire
                did_change = True

            # last_salary_adjustment_date：merge 与 replace 都保留旧值（时间戳无 NULL 清空语义）
            if new_last_adj is not None and new_last_adj != old_last_adj:
                employee.last_salary_adjustment_date = new_last_adj
                did_change = True

            results.append({
                'row_index': row_no,
                'row': row_no,
                'status': 'success',
                'employee_no': emp_no,
                'action': 'update' if did_change else 'no_change',
                'message': '入职信息导入成功。' if did_change else '无字段变化，跳过更新。',
                'fields': {
                    'hire_date': {
                        'old': old_hire.isoformat() if old_hire else None,
                        'new': new_hire.isoformat() if new_hire else None,
                    },
                    'last_salary_adjustment_date': {
                        'old': old_last_adj.isoformat() if old_last_adj else None,
                        'new': new_last_adj.isoformat() if new_last_adj else None,
                    },
                },
            })

        self.db.flush()
        self.db.commit()
        return results

    # ------------------------------------------------------------------
    # Phase 32 D-03: Non-statutory leave import (IMPORT-01)
    # ------------------------------------------------------------------

    def _import_non_statutory_leave(
        self,
        dataframe: pd.DataFrame,
        *,
        overwrite_mode: str = 'merge',  # Phase 32 IMPORT-05
    ) -> list[dict[str, object]]:
        """D-03 + IMPORT-05：导入非法定假期到 NonStatutoryLeave 表，业务键 = (employee_id, year)。

        upsert 语义：
          - 第一次见到 (employee_id, year)：insert
          - 已存在：merge 空字段保留旧值；replace 可选字段 leave_type 空时清空、必填字段
            total_days 空时 row failed
        """
        # 业务键定义在类常量 _BUSINESS_KEYS（Phase 32-03 已抽取）
        # preview 阶段同文件冲突检测复用同一份 key 列表
        results: list[dict[str, object]] = []

        employee_nos = [
            str(v).strip()
            for v in dataframe.get('employee_no', [])
            if v is not None and str(v).strip()
        ]
        emp_rows = (
            self.db.execute(
                select(Employee).where(Employee.employee_no.in_(employee_nos))
            ).scalars().all()
            if employee_nos else []
        )
        emp_map: dict[str, Employee] = {e.employee_no: e for e in emp_rows}

        for index, row in dataframe.iterrows():
            row_no = int(index) + 1
            emp_no = str(row.get('employee_no', '')).strip()
            if not emp_no:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': '',
                    'error': '员工工号为必填字段',
                    'message': '员工工号为必填字段',
                    'error_column': '员工工号',
                })
                continue

            employee = emp_map.get(emp_no)
            if employee is None:
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': f'未找到员工工号 "{emp_no}"',
                    'message': f'未找到员工工号 "{emp_no}"',
                    'error_column': '员工工号',
                })
                continue

            # year: 三层转换防 '2026.0' 等浮点字符串
            raw_year = row.get('year')
            try:
                year = int(float(str(raw_year).strip()))
            except (ValueError, TypeError):
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'failed',
                    'employee_no': emp_no,
                    'error': f'年度字段无效: {raw_year!r}',
                    'message': f'年度字段无效: {raw_year!r}',
                    'error_column': '年度',
                })
                continue

            # total_days: Decimal 精度（Pitfall 8）
            raw_days = row.get('total_days')
            total_days_decimal: Decimal | None = None
            if raw_days is None or str(raw_days).strip() == '':
                if overwrite_mode == 'replace':
                    results.append({
                        'row_index': row_no,
                        'row': row_no,
                        'status': 'failed',
                        'employee_no': emp_no,
                        'error': '替换模式下假期天数为必填字段',
                        'message': '替换模式下假期天数为必填字段',
                        'error_column': '假期天数',
                    })
                    continue
            else:
                try:
                    total_days_decimal = Decimal(str(raw_days).strip())
                except InvalidOperation:
                    results.append({
                        'row_index': row_no,
                        'row': row_no,
                        'status': 'failed',
                        'employee_no': emp_no,
                        'error': f'假期天数格式无效: {raw_days!r}',
                        'message': f'假期天数格式无效: {raw_days!r}',
                        'error_column': '假期天数',
                    })
                    continue

            raw_leave_type = row.get('leave_type') if 'leave_type' in dataframe.columns else None
            leave_type: str | None
            if raw_leave_type is None or str(raw_leave_type).strip() == '':
                leave_type = None
            else:
                leave_type = str(raw_leave_type).strip()

            existing = self.db.execute(
                select(NonStatutoryLeave).where(
                    NonStatutoryLeave.employee_id == employee.id,
                    NonStatutoryLeave.year == year,
                )
            ).scalar_one_or_none()

            if existing is None:
                if total_days_decimal is None:
                    # 新增记录时假期天数必须有值
                    results.append({
                        'row_index': row_no,
                        'row': row_no,
                        'status': 'failed',
                        'employee_no': emp_no,
                        'error': '新增记录时假期天数为必填',
                        'message': '新增记录时假期天数为必填',
                        'error_column': '假期天数',
                    })
                    continue
                record = NonStatutoryLeave(
                    employee_id=employee.id,
                    employee_no=emp_no,
                    year=year,
                    total_days=total_days_decimal,
                    leave_type=leave_type,
                    source='excel',
                )
                self.db.add(record)
                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'success',
                    'employee_no': emp_no,
                    'action': 'insert',
                    'message': '非法定假期记录新增成功。',
                    'fields': {
                        'year': {'old': None, 'new': year},
                        'total_days': {'old': None, 'new': str(total_days_decimal)},
                        'leave_type': {'old': None, 'new': leave_type},
                    },
                })
            else:
                old_days = existing.total_days
                old_type = existing.leave_type
                did_change = False

                # total_days：必填字段，merge 空保留 / replace 空已在上方拦截
                if total_days_decimal is not None and total_days_decimal != old_days:
                    existing.total_days = total_days_decimal
                    did_change = True

                # leave_type：可选字段
                #   merge：空保留旧值
                #   replace：空清空旧值（如果有的话）
                if leave_type is None:
                    if overwrite_mode == 'replace' and old_type is not None:
                        existing.leave_type = None
                        did_change = True
                else:
                    if leave_type != old_type:
                        existing.leave_type = leave_type
                        did_change = True

                results.append({
                    'row_index': row_no,
                    'row': row_no,
                    'status': 'success',
                    'employee_no': emp_no,
                    'action': 'update' if did_change else 'no_change',
                    'message': '非法定假期记录更新成功。' if did_change else '无字段变化，跳过更新。',
                    'fields': {
                        'total_days': {
                            'old': str(old_days) if old_days is not None else None,
                            'new': str(total_days_decimal) if total_days_decimal is not None else str(old_days),
                        },
                        'leave_type': {'old': old_type, 'new': leave_type},
                    },
                })

        self.db.flush()
        self.db.commit()
        return results

    # ==================================================================
    # Phase 32-03 D-06 / D-07 / D-08 / D-09: 两阶段提交之 preview 阶段
    # ==================================================================

    def _detect_in_file_conflicts(
        self,
        import_type: str,
        dataframe: pd.DataFrame,
    ) -> dict[int, str]:
        """D-09: 同一批文件内业务键重复检测。

        按 _BUSINESS_KEYS 中定义的业务键分组，count > 1 的 group 内所有行
        都返回 conflict reason。pandas dropna=False 保证 NaN 也参与分组
        （避免空值被分组忽略）。

        Returns: {pandas_idx: conflict_reason} 仅含冲突行
        """
        keys = self._BUSINESS_KEYS.get(import_type)
        if not keys or not all(k in dataframe.columns for k in keys):
            return {}
        grouped = dataframe.groupby(keys, dropna=False)
        conflicts: dict[int, str] = {}
        for key_tuple, group in grouped:
            if len(group) > 1:
                if not isinstance(key_tuple, tuple):
                    key_tuple = (key_tuple,)
                key_str = ', '.join(
                    f'{k}={v}' for k, v in zip(keys, key_tuple)
                )
                reason = (
                    f'同文件内 ({key_str}) 出现 {len(group)} 次，'
                    f'请仅保留一行后重新上传'
                )
                for idx in group.index:
                    conflicts[int(idx)] = reason
        return conflicts

    def _build_row_diff(
        self,
        import_type: str,
        row: 'pd.Series',
    ) -> tuple[str, dict[str, dict]]:
        """D-08: 逐行算 diff，返回 (action, fields_dict)。

        简化策略：仅查 DB 现状对比新值，**不真正落库**（preview 是只读阶段）。

        各 import_type 的查询逻辑：
          - hire_info: 找 Employee by employee_no，对比 hire_date / last_salary_adjustment_date
          - non_statutory_leave: 找 NonStatutoryLeave by (employee_id, year)
          - performance_grades: 找 PerformanceRecord by (employee_id, year)
          - salary_adjustments: 找 SalaryAdjustmentRecord by (employee_id, adjustment_date, adjustment_type)

        action 取值：'insert' | 'update' | 'no_change' | 'conflict'
        ('conflict' 仅在员工不存在或字段解析失败时返回；同文件冲突由
        _detect_in_file_conflicts 提前拦截，不会进到本函数。)
        """
        emp_no = str(row.get('employee_no', '')).strip()
        if not emp_no:
            return 'no_change', {}

        emp = self.db.execute(
            select(Employee).where(Employee.employee_no == emp_no)
        ).scalar_one_or_none()
        if emp is None:
            return 'conflict', {}

        if import_type == 'hire_info':
            try:
                new_hire = self._parse_excel_date(row.get('hire_date'))
                new_last_adj = self._parse_excel_date(
                    row.get('last_salary_adjustment_date')
                )
            except ValueError:
                return 'conflict', {}
            old_hire = emp.hire_date
            old_last_adj = emp.last_salary_adjustment_date
            fields: dict[str, dict] = {}
            changed = False
            if new_hire is not None and new_hire != old_hire:
                fields['hire_date'] = {
                    'old': old_hire.isoformat() if old_hire else None,
                    'new': new_hire.isoformat(),
                }
                changed = True
            if new_last_adj is not None and new_last_adj != old_last_adj:
                fields['last_salary_adjustment_date'] = {
                    'old': old_last_adj.isoformat() if old_last_adj else None,
                    'new': new_last_adj.isoformat(),
                }
                changed = True
            if not changed:
                return 'no_change', {}
            return 'update', fields

        if import_type == 'non_statutory_leave':
            try:
                year = int(float(str(row.get('year', '')).strip()))
            except (ValueError, TypeError):
                return 'conflict', {}
            existing = self.db.execute(
                select(NonStatutoryLeave).where(
                    NonStatutoryLeave.employee_id == emp.id,
                    NonStatutoryLeave.year == year,
                )
            ).scalar_one_or_none()
            new_days_str = str(row.get('total_days', '')).strip()
            raw_type = row.get('leave_type')
            new_type = (
                str(raw_type).strip()
                if raw_type is not None and str(raw_type).strip()
                else None
            )
            if existing is None:
                return 'insert', {
                    'year': {'old': None, 'new': year},
                    'total_days': {'old': None, 'new': new_days_str or None},
                    'leave_type': {'old': None, 'new': new_type},
                }
            try:
                new_days = Decimal(new_days_str) if new_days_str else None
            except InvalidOperation:
                return 'conflict', {}
            fields = {}
            if new_days is not None and new_days != existing.total_days:
                fields['total_days'] = {
                    'old': str(existing.total_days),
                    'new': str(new_days),
                }
            if new_type is not None and new_type != existing.leave_type:
                fields['leave_type'] = {
                    'old': existing.leave_type,
                    'new': new_type,
                }
            return ('update' if fields else 'no_change'), fields

        if import_type == 'performance_grades':
            try:
                year = int(float(str(row.get('year', '')).strip()))
            except (ValueError, TypeError):
                return 'conflict', {}
            existing = self.db.execute(
                select(PerformanceRecord).where(
                    PerformanceRecord.employee_id == emp.id,
                    PerformanceRecord.year == year,
                )
            ).scalar_one_or_none()
            new_grade = str(row.get('grade', '')).strip()
            if existing is None:
                return 'insert', {'grade': {'old': None, 'new': new_grade}}
            if existing.grade == new_grade:
                return 'no_change', {}
            return 'update', {'grade': {'old': existing.grade, 'new': new_grade}}

        if import_type == 'salary_adjustments':
            try:
                adj_date = self._parse_excel_date(row.get('adjustment_date'))
            except ValueError:
                return 'conflict', {}
            raw_type = row.get('adjustment_type')
            adj_type = (
                str(raw_type).strip()
                if raw_type is not None and str(raw_type).strip()
                else None
            )
            if not adj_date or not adj_type:
                return 'conflict', {}
            # 与 _import_salary_adjustments 一致，做中文标签映射
            adj_type_code = self._ADJ_TYPE_MAP.get(adj_type, adj_type.lower())
            existing = self.db.execute(
                select(SalaryAdjustmentRecord).where(
                    SalaryAdjustmentRecord.employee_id == emp.id,
                    SalaryAdjustmentRecord.adjustment_date == adj_date,
                    SalaryAdjustmentRecord.adjustment_type == adj_type_code,
                )
            ).scalar_one_or_none()
            new_amount_str = str(row.get('amount', '')).strip()
            if existing is None:
                return 'insert', {
                    'amount': {
                        'old': None,
                        'new': new_amount_str or None,
                    }
                }
            old_amount_str = (
                str(existing.amount) if existing.amount is not None else None
            )
            if old_amount_str == (new_amount_str or None):
                return 'no_change', {}
            return 'update', {
                'amount': {
                    'old': old_amount_str,
                    'new': new_amount_str or None,
                }
            }

        return 'no_change', {}

    def build_preview(
        self,
        *,
        import_type: str,
        file_name: str,
        raw_bytes: bytes,
        actor_id: str | None = None,
    ) -> 'PreviewResponse':  # type: ignore[name-defined]
        """D-06 + D-07 + D-08 + D-09: preview 阶段（两阶段提交之第一步）。

        流程：
          1) 解析 xlsx → dataframe
          2) 创建 ImportJob status='previewing' + 落盘暂存文件 + 算 sha256
          3) 检测同文件内业务键冲突
          4) 逐行算 diff（insert/update/no_change/conflict）
          5) 按优先级截断到 200 行（conflict > insert > update > no_change）
          6) 写 result_summary.preview = {counters, rows, file_sha256, expires_at}
          7) 返回 PreviewResponse

        Raises:
            ValueError: import_type 不支持 / 文件为空 / 行数超限
        """
        # 延迟 import 避免循环：schemas/import_preview.py 仅 service 用
        from backend.app.schemas.import_preview import (
            FieldDiff,
            PreviewCounters,
            PreviewResponse,
            PreviewRow,
        )

        normalized_type = import_type.strip().lower()
        if normalized_type not in self.SUPPORTED_TYPES:
            raise ValueError(
                self._localize_error_message('Unsupported import type.')
            )
        if not raw_bytes:
            raise ValueError(
                self._localize_error_message('Uploaded file is empty.')
            )

        # 解析 xlsx
        dataframe = self._load_table(file_name, raw_bytes)
        if len(dataframe) > self.MAX_ROWS:
            raise ValueError(
                f'单次导入不能超过 {self.MAX_ROWS} 行，请分批导入。'
            )
        dataframe = self._normalize_columns(normalized_type, dataframe)

        # 创建 ImportJob status='previewing'（默认 overwrite_mode='merge'，
        # confirm 阶段按用户选择更新）
        job = ImportJob(
            file_name=file_name,
            import_type=normalized_type,
            status='previewing',
            total_rows=len(dataframe),
            success_rows=0,
            failed_rows=0,
            result_summary={},
            overwrite_mode='merge',
            actor_id=actor_id,
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)

        # 暂存文件 + 算 sha256（防 confirm 阶段被外部篡改 / T-32-14）
        file_sha256 = self._save_staged_file(job.id, raw_bytes)

        # 检测同文件冲突
        conflicts = self._detect_in_file_conflicts(normalized_type, dataframe)

        rows_all: list[PreviewRow] = []
        counters = {
            'insert': 0, 'update': 0, 'no_change': 0, 'conflict': 0,
        }

        for pandas_idx, row in dataframe.iterrows():
            row_no = int(pandas_idx) + 2  # +2: pandas idx 0 = Excel 第二行（含表头）
            emp_no = str(row.get('employee_no', '')).strip()

            if pandas_idx in conflicts:
                rows_all.append(PreviewRow(
                    row_number=row_no,
                    action='conflict',
                    employee_no=emp_no,
                    fields={},
                    conflict_reason=conflicts[pandas_idx],
                ))
                counters['conflict'] += 1
                continue

            try:
                action, fields = self._build_row_diff(normalized_type, row)
            except Exception as exc:
                rows_all.append(PreviewRow(
                    row_number=row_no,
                    action='conflict',
                    employee_no=emp_no,
                    fields={},
                    conflict_reason=f'行解析失败: {exc}',
                ))
                counters['conflict'] += 1
                continue

            counters[action] += 1
            rows_all.append(PreviewRow(
                row_number=row_no,
                action=action,
                employee_no=emp_no,
                fields={k: FieldDiff(**v) for k, v in fields.items()},
            ))

        # 截断策略（D-08）：conflict > insert > update > no_change 优先级保留
        priority = {'conflict': 0, 'insert': 1, 'update': 2, 'no_change': 3}
        rows_sorted = sorted(rows_all, key=lambda r: priority[r.action])
        rows_kept = rows_sorted[:200]
        rows_truncated = len(rows_all) > 200

        expires_at = datetime.now(timezone.utc) + timedelta(minutes=60)

        # 把 preview 结果存入 ImportJob.result_summary（confirm 时只更新 execution 部分）
        job.result_summary = {
            'preview': {
                'counters': counters,
                'rows': [r.model_dump() for r in rows_kept],
                'rows_truncated': rows_truncated,
                'truncated_count': max(0, len(rows_all) - 200),
                'file_sha256': file_sha256,
                'preview_expires_at': expires_at.isoformat(),
            },
            'supported_types': sorted(self.SUPPORTED_TYPES),
        }
        self.db.commit()
        self.db.refresh(job)

        return PreviewResponse(
            job_id=job.id,
            import_type=normalized_type,
            file_name=file_name,
            total_rows=len(dataframe),
            counters=PreviewCounters(**counters),
            rows=rows_kept,
            rows_truncated=rows_truncated,
            truncated_count=max(0, len(rows_all) - 200),
            preview_expires_at=expires_at,
            file_sha256=file_sha256,
        )

    # ==================================================================
    # Phase 32-03 D-06 / D-13: 两阶段提交之 confirm + cancel 阶段
    # ==================================================================

    def confirm_import(
        self,
        *,
        job_id: str,
        overwrite_mode: str,
        confirm_replace: bool = False,
        actor_id: str | None = None,
        actor_role: str | None = None,
    ) -> 'ConfirmResponse':  # type: ignore[name-defined]
        """D-06 + D-13: confirm 阶段（两阶段提交之第二步）。

        流程：
          1) 校验 overwrite_mode（merge / replace） 与 replace 二次确认
             （T-32-15: replace 必须 confirm_replace=True）
          2) 校验 job 状态必须是 'previewing'（防双 confirm）
          3) 读暂存文件 + sha256 校验（T-32-14: 防外部篡改）
          4) 标记 job.status='processing' 并写入最终 overwrite_mode
          5) 调 _dispatch_import 落库 → 派生 status (completed/partial/failed)
          6) 删暂存文件
          7) 写 AuditLog action='import_confirmed'，用真实字段名
             operator_id/target_type/target_id（不是 D-13 文档假设的 actor_id/resource_*）

        Raises:
            ValueError: overwrite_mode 不合法 / replace 未确认 / job 不存在
                / job 状态非 previewing / 暂存文件 hash 不匹配
        """
        from backend.app.schemas.import_preview import ConfirmResponse

        if overwrite_mode not in ('merge', 'replace'):
            raise ValueError(f'Invalid overwrite_mode: {overwrite_mode!r}')
        if overwrite_mode == 'replace' and not confirm_replace:
            raise ValueError('替换模式必须显式确认（前端二次确认未完成）')

        job = self.db.execute(
            select(ImportJob).where(ImportJob.id == job_id)
        ).scalar_one_or_none()
        if job is None:
            raise ValueError(f'ImportJob {job_id} 不存在')
        if job.status != 'previewing':
            raise ValueError(
                f'ImportJob {job_id} 状态为 {job.status}，'
                f'无法确认导入（已确认或已取消）'
            )

        expected_sha = (
            (job.result_summary or {}).get('preview', {}).get('file_sha256')
        )
        raw_bytes = self._read_staged_file(job_id, expected_sha256=expected_sha)

        # 标记 processing + 落 overwrite_mode
        job.status = 'processing'
        job.overwrite_mode = overwrite_mode
        if actor_id is not None:
            job.actor_id = actor_id
        self.db.commit()

        start_ts = datetime.now(timezone.utc)
        inserted = updated = no_change = failed = 0
        duration_ms = 0
        try:
            dataframe = self._load_table(job.file_name, raw_bytes)
            dataframe = self._normalize_columns(job.import_type, dataframe)
            row_results = self._dispatch_import(
                job.import_type, dataframe, overwrite_mode=overwrite_mode,
            )
            inserted = sum(
                1 for r in row_results
                if r.get('action') == 'insert' and r.get('status') == 'success'
            )
            updated = sum(
                1 for r in row_results
                if r.get('action') == 'update' and r.get('status') == 'success'
            )
            no_change = sum(
                1 for r in row_results
                if r.get('action') == 'no_change' and r.get('status') == 'success'
            )
            failed = sum(
                1 for r in row_results if r.get('status') == 'failed'
            )

            job.total_rows = len(row_results)
            job.success_rows = inserted + updated + no_change
            job.failed_rows = failed
            if failed == 0:
                job.status = 'completed'
            elif job.success_rows == 0:
                job.status = 'failed'
            else:
                job.status = 'partial'
            duration_ms = int(
                (datetime.now(timezone.utc) - start_ts).total_seconds() * 1000
            )
            existing = job.result_summary or {}
            job.result_summary = {
                **existing,
                'execution': {
                    'rows': row_results[:200],  # 截断保护，与 preview 同口径
                    'inserted_count': inserted,
                    'updated_count': updated,
                    'no_change_count': no_change,
                    'failed_count': failed,
                    'execution_duration_ms': duration_ms,
                    'executed_at': datetime.now(timezone.utc).isoformat(),
                },
            }
        except Exception as exc:
            job.status = 'failed'
            existing = job.result_summary or {}
            job.result_summary = {
                **existing, 'execution': {'error': str(exc)},
            }
            self.db.commit()
            raise
        finally:
            try:
                self._delete_staged_file(job_id)
            except Exception:
                logger.exception(
                    'Failed to delete staged file for confirmed job %s', job_id,
                )

        # D-13: 写 AuditLog（**真实字段名 operator_id / target_type / target_id**）
        # 注意：AuditLog model 字段是 operator_id/target_type/target_id，
        # 与 D-13 文档原写的 actor_id/resource_type/resource_id 不一致 —— 用真实字段
        audit = AuditLog(
            operator_id=actor_id,
            operator_role=actor_role,
            action='import_confirmed',
            target_type='import_job',
            target_id=job.id,
            detail={
                'import_type': job.import_type,
                'overwrite_mode': overwrite_mode,
                'file_name': job.file_name,
                'total_rows': job.total_rows,
                'inserted_count': inserted,
                'updated_count': updated,
                'no_change_count': no_change,
                'failed_count': failed,
            },
        )
        self.db.add(audit)
        self.db.commit()
        self.db.refresh(job)

        # Phase 34 D-03 / D-15 / W-1：performance_grades 导入成功后同步触发档次重算
        # 最长阻塞 5 秒；超时后台继续；失败不阻塞 import 落库（D-04）
        tier_recompute_status: str | None = None
        if (
            job.import_type == 'performance_grades'
            and job.status in ('completed', 'partial')
        ):
            tier_recompute_status = self._run_tier_recompute_hook(job)

        return ConfirmResponse(
            job_id=job.id,
            status=job.status,
            total_rows=job.total_rows,
            inserted_count=inserted,
            updated_count=updated,
            no_change_count=no_change,
            failed_count=failed,
            execution_duration_ms=duration_ms,
            tier_recompute_status=tier_recompute_status,
        )

    def _run_tier_recompute_hook(self, job: ImportJob) -> str:
        """Phase 34 D-03 / D-04 / D-06：performance_grades confirm 同步重算 hook。

        返回 ConfirmResponse.tier_recompute_status 字段值：
          - 'completed'    — 5 秒内全部 year 重算成功
          - 'in_progress'  — 5 秒超时；后台线程仍在重算
          - 'busy_skipped' — 至少一个 year 撞上 HR 手动重算锁（D-06）
          - 'failed'       — 至少一个 year 重算异常（D-04 已落库不阻塞）
        """
        # 延迟 import 避免循环依赖
        from concurrent.futures import (
            ThreadPoolExecutor,
            TimeoutError as FutureTimeout,
        )
        from datetime import timedelta

        from backend.app.core.config import get_settings as _get_settings
        from backend.app.core.redis import get_redis as _get_redis
        from backend.app.services.exceptions import (
            TierRecomputeBusyError,
            TierRecomputeFailedError,
        )
        from backend.app.services.performance_service import PerformanceService
        from backend.app.services.tier_cache import TierCache

        # 抽取受影响 years
        affected_years: set[int] = set()
        exec_rows = (
            (job.result_summary or {}).get('execution', {}).get('rows', [])
        )
        for row in exec_rows:
            year_val = row.get('year') if isinstance(row, dict) else None
            if year_val is not None:
                try:
                    affected_years.add(int(year_val))
                except (TypeError, ValueError):
                    pass

        if not affected_years:
            # 兜底：rows 不含 year 时查 source='excel' 最近 1 分钟内的不同 year
            one_minute_ago = datetime.now(timezone.utc) - timedelta(minutes=1)
            year_rows = self.db.execute(
                select(PerformanceRecord.year).where(
                    PerformanceRecord.source == 'excel',
                    PerformanceRecord.updated_at >= one_minute_ago,
                ).distinct()
            ).scalars().all()
            affected_years = {int(y) for y in year_rows if y is not None}

        if not affected_years:
            logger.info(
                'No affected years detected for performance_grades job %s; '
                'skip tier recompute',
                job.id,
            )
            return 'completed'

        settings = _get_settings()
        try:
            redis_client = _get_redis()
        except Exception:  # noqa: BLE001 — Redis 不可达走 None 降级
            redis_client = None
        cache = TierCache(redis_client=redis_client, settings=settings)
        perf_service = PerformanceService(
            self.db, settings=settings, cache=cache,
        )

        # invalidate cache for all affected years
        try:
            perf_service.invalidate_tier_cache(affected_years)
        except Exception:  # noqa: BLE001
            logger.exception(
                'Tier cache invalidate failed for years %s', affected_years,
            )

        timeout_s = settings.performance_tier_recompute_timeout_seconds
        busy_seen = False
        failed_seen = False

        def _do() -> None:
            nonlocal busy_seen, failed_seen
            for y in sorted(affected_years):
                try:
                    perf_service.recompute_tiers(y)
                except TierRecomputeBusyError:
                    busy_seen = True
                    logger.warning(
                        'Tier recompute busy for year %s '
                        '(D-06 self-vs-manual collision)',
                        y,
                    )
                except TierRecomputeFailedError as fe:
                    failed_seen = True
                    logger.warning(
                        'Tier recompute failed for year %s: %s', y, fe.cause,
                    )

        try:
            # ThreadPoolExecutor 不强制 shutdown 让超时后后台继续运行
            ex = ThreadPoolExecutor(max_workers=1)
            try:
                fut = ex.submit(_do)
                try:
                    fut.result(timeout=timeout_s)
                    if failed_seen:
                        return 'failed'
                    if busy_seen:
                        return 'busy_skipped'
                    return 'completed'
                except FutureTimeout:
                    # 超时：后台线程继续运行，executor 不强制 shutdown
                    return 'in_progress'
            finally:
                # wait=False：不等待后台任务（D-03 让其继续完成）
                ex.shutdown(wait=False)
        except Exception as exc:  # noqa: BLE001
            logger.exception('Tier recompute hook fatal error: %s', exc)
            return 'failed'

    def cancel_import(self, job_id: str) -> None:
        """HR 显式取消 preview → status='cancelled' + 删暂存文件。

        对终态 job（completed/failed/partial/cancelled）幂等：直接返回，
        不抛异常，不改状态。
        """
        job = self.db.execute(
            select(ImportJob).where(ImportJob.id == job_id)
        ).scalar_one_or_none()
        if job is None:
            raise ValueError(f'ImportJob {job_id} 不存在')
        if job.status != 'previewing':
            # 只能取消 previewing 状态；其他状态忽略（幂等，不改）
            return
        job.status = 'cancelled'
        existing = job.result_summary or {}
        job.result_summary = {
            **existing, 'cancellation_reason': 'user_cancelled',
        }
        try:
            self._delete_staged_file(job_id)
        except Exception:
            logger.exception(
                'Failed to delete staged file for cancelled job %s', job_id,
            )
        self.db.commit()

    # ==================================================================
    # Phase 32-03 D-16 / D-17: 并发锁 + 双时限僵尸清理
    # ==================================================================

    def is_import_running(self, import_type: str | None = None) -> bool:
        """D-16: 检测同 import_type 是否有活跃 job（previewing 或 processing 状态）。

        per-import_type 分桶锁：传 import_type 时只查该 type 的活跃 job
        （4 类资格 import_type 互不影响，可并行）；不传时全局检查。
        """
        stmt = select(ImportJob).where(
            ImportJob.status.in_(list(self._LOCKING_STATUSES))
        )
        if import_type is not None:
            stmt = stmt.where(ImportJob.import_type == import_type)
        return self.db.execute(stmt.limit(1)).scalar_one_or_none() is not None

    def get_active_job(self, import_type: str) -> ImportJob | None:
        """D-18: 取该 import_type 的活跃 job（最多一条；多条时返回最新创建的）。"""
        stmt = (
            select(ImportJob)
            .where(
                ImportJob.status.in_(list(self._LOCKING_STATUSES)),
                ImportJob.import_type == import_type,
            )
            .order_by(ImportJob.created_at.desc())
            .limit(1)
        )
        return self.db.execute(stmt).scalar_one_or_none()

    def expire_stale_import_jobs(
        self,
        *,
        processing_timeout_minutes: int = 30,
        previewing_timeout_minutes: int = 60,
    ) -> dict[str, int]:
        """D-17: 双时限僵尸清理。

        - processing 超过 30 分钟 → 'failed'，写 result_summary.error='timeout'
        - previewing 超过 60 分钟 → 'cancelled'，删除暂存文件并写 cancellation_reason

        终态（completed/failed/partial/cancelled）保持不动。

        Returns: {'processing': N, 'previewing': M} 实际清理数量
        """
        now = datetime.now(timezone.utc)
        expired = {'processing': 0, 'previewing': 0}

        # processing 30min → failed
        cutoff_p = now - timedelta(minutes=processing_timeout_minutes)
        stale_p = list(
            self.db.execute(
                select(ImportJob).where(
                    ImportJob.status == 'processing',
                    ImportJob.created_at < cutoff_p,
                )
            ).scalars().all()
        )
        for job in stale_p:
            job.status = 'failed'
            existing = job.result_summary or {}
            job.result_summary = {
                **existing,
                'error': 'timeout',
                'expired_at': now.isoformat(),
            }
        expired['processing'] = len(stale_p)

        # previewing 60min → cancelled + 删暂存文件
        cutoff_v = now - timedelta(minutes=previewing_timeout_minutes)
        stale_v = list(
            self.db.execute(
                select(ImportJob).where(
                    ImportJob.status == 'previewing',
                    ImportJob.created_at < cutoff_v,
                )
            ).scalars().all()
        )
        for job in stale_v:
            job.status = 'cancelled'
            existing = job.result_summary or {}
            job.result_summary = {
                **existing,
                'cancellation_reason': 'preview_timeout',
                'cancelled_at': now.isoformat(),
            }
            try:
                self._delete_staged_file(job.id)
            except Exception:
                logger.exception(
                    'Failed to delete staged file for expired job %s', job.id,
                )
        expired['previewing'] = len(stale_v)

        if stale_p or stale_v:
            self.db.commit()
            logger.warning(
                'expire_stale_import_jobs cleaned up %d processing + %d previewing',
                expired['processing'], expired['previewing'],
            )
        return expired

    # ==================================================================
    # Phase 32-03 D-06: 暂存文件管理（含 T-32-01 路径遍历防护）
    # ==================================================================

    @staticmethod
    def _staged_file_path(job_id: str) -> Path:
        """安全的暂存文件路径，job_id 必须是 UUID（无路径分隔符与 .. 段）。

        T-32-01 双重防护（参考 backend/app/core/storage.py LocalStorageService）：
        1. 字符级校验：拒绝空、'/', '\\', '..'
        2. resolve 后 is_relative_to(base_dir) 二次校验

        Returns: <storage_base_dir>/imports/<job_id>.xlsx 的绝对路径
        Raises: ValueError 当 job_id 不安全
        """
        if not job_id or '/' in job_id or '\\' in job_id or '..' in job_id:
            raise ValueError(
                f'Invalid job_id (path traversal blocked): {job_id!r}'
            )
        base = (Path(get_settings().storage_base_dir).resolve() / 'imports').resolve()
        base.mkdir(parents=True, exist_ok=True)
        target = (base / f'{job_id}.xlsx').resolve()
        if not target.is_relative_to(base):
            # 字符级 + resolve 双重防护后理论上不会到这里；保底再校验一次
            raise ValueError(
                f'Staged file path traversal escape detected: {job_id!r}'
            )
        return target

    def _save_staged_file(self, job_id: str, content: bytes) -> str:
        """存盘 + 返回 sha256（preview 阶段写入 result_summary.preview.file_sha256）。"""
        path = self._staged_file_path(job_id)
        path.write_bytes(content)
        return hashlib.sha256(content).hexdigest()

    def _read_staged_file(
        self,
        job_id: str,
        *,
        expected_sha256: str | None = None,
    ) -> bytes:
        """读取暂存文件，可选校验 hash（T-32-14: confirm 阶段防外部篡改）。

        Raises:
            ValueError: 文件不存在 / hash 不匹配
        """
        path = self._staged_file_path(job_id)
        if not path.exists():
            raise ValueError(
                f'Staged file for job {job_id} not found (expired or deleted)'
            )
        content = path.read_bytes()
        if expected_sha256 is not None:
            actual = hashlib.sha256(content).hexdigest()
            if actual != expected_sha256:
                raise ValueError(
                    f'Staged file hash mismatch for job {job_id}: '
                    f'expected {expected_sha256[:8]}..., got {actual[:8]}... '
                    f'(file modified externally; please re-upload)'
                )
        return content

    def _delete_staged_file(self, job_id: str) -> None:
        """删除暂存文件（不存在不报错）。"""
        path = self._staged_file_path(job_id)
        if path.exists():
            path.unlink()
